from django.core.management.base import BaseCommand
from django.contrib.contenttypes.models import ContentType
from django.contrib.auth import get_user_model
from django.db import transaction
from lab.models import (
    Individual,
    Status,
    IdentifierType,
    Sample,
    SampleType,
    Test,
    TestType,
    Analysis,
    AnalysisType,
)
import openpyxl
from datetime import datetime


User = get_user_model()


class Command(BaseCommand):
    help = "Update individuals from 'Yayın_İçi_Tablo_Özbek_Lab.xlsx' sheet 'yayıniciyedek'"

    def add_arguments(self, parser):
        parser.add_argument(
            "file_path",
            type=str,
            nargs="?",
            default="Yayın_İçi_Tablo_Özbek_Lab.xlsx",
            help="Path to the XLSX file (defaults to 'Yayın_İçi_Tablo_Özbek_Lab.xlsx' in CWD)",
        )
        parser.add_argument(
            "--admin-username",
            type=str,
            required=True,
            help="Admin username used as created_by for status creation",
        )

    def _parse_ddmmyyyy(self, value):
        if not value:
            return None
        # Excel may provide datetime/date types directly
        if hasattr(value, "date"):
            try:
                return value.date()
            except Exception:
                return None
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
        return None

    def _normalize_sex(self, value):
        if value is None:
            return None
        text = str(value).strip().lower()
        if text in {"m", "male", "erkek"}:
            return "male"
        if text in {"f", "female", "kadin", "kadın"}:
            return "female"
        if text in {"other", "o"}:
            return "other"
        return None

    def _to_bool(self, value):
        # Accept 1/0, True/False, "1"/"0"
        if value in (1, True, "1", "True", "true"):
            return True
        if value in (0, False, "0", "False", "false"):
            return False
        return None

    def _ensure_statuses(self, admin_user):
        ct_individual = ContentType.objects.get(app_label="lab", model="individual")
        # name, color, icon
        desired = [
            ("Unsolved", "red", "fa-circle-cross"),
            ("Solved - P/LP", "green", "fa-circle-check"),
            ("Solved - VUS", "lightgreen", "fa-circle-check"),
            ("Novel Gene Disease Assoc.", "purple", "fa-plus"),
        ]
        for name, color, icon in desired:
            status_obj, created = Status.objects.get_or_create(
                name=name,
                content_type=ct_individual,
                defaults={
                    "description": "",
                    "color": color,
                    "created_by": admin_user,
                    "icon": icon,
                },
            )
            # Backfill icon if already exists but missing/different
            changed = False
            if not created and icon and status_obj.icon != icon:
                status_obj.icon = icon
                changed = True
            # Backfill color if different casing/value was used previously
            if status_obj.color != color:
                status_obj.color = color
                changed = True
            if changed:
                status_obj.save()

        # Ensure Test status "Previous" exists
        ct_test = ContentType.objects.get(app_label="lab", model="test")
        prev_status, created = Status.objects.get_or_create(
            name="Previous",
            content_type=ct_test,
            defaults={
                "description": "Historical test performed before RareBoost",
                "color": "orange",
                "created_by": admin_user,
                "icon": "fa-clock-rotate-left",
            },
        )
        if not created:
            changed = False
            if prev_status.color != "orange":
                prev_status.color = "orange"
                changed = True
            if prev_status.icon != "fa-clock-rotate-left":
                prev_status.icon = "fa-clock-rotate-left"
                changed = True
            if changed:
                prev_status.save()

    def _get_or_create_test_type(self, name, admin_user):
        if not name:
            return None
        tt, _ = TestType.objects.get_or_create(name=name, defaults={"created_by": admin_user})
        return tt

    def _get_or_create_analysis_type(self, name, admin_user):
        if not name:
            return None
        at, _ = AnalysisType.objects.get_or_create(name=name, defaults={"created_by": admin_user})
        return at

    def _get_or_create_sample_type(self, name, admin_user):
        st, _ = SampleType.objects.get_or_create(name=name, defaults={"created_by": admin_user})
        return st

    def _ensure_sample_for_tests(self, individual, admin_user):
        # Prefer existing Whole Blood sample
        whole_blood = SampleType.objects.filter(name__iexact="Whole Blood").first()
        if whole_blood:
            sample = individual.samples.filter(sample_type=whole_blood).first()
            if sample:
                return sample
        # If not existing, create placeholder sample
        placeholder_type = self._get_or_create_sample_type("Placeholder", admin_user)
        ct_sample = ContentType.objects.get(app_label="lab", model="sample")
        not_available_status, _ = Status.objects.get_or_create(
            name="Not Available",
            content_type=ct_sample,
            defaults={
                "description": "A placeholder sample for tests performed off-center",
                "color": "gray",
                "created_by": admin_user,
                "icon": "fa-ban",
            },
        )
        sample = Sample.objects.create(
            individual=individual,
            sample_type=placeholder_type,
            status=not_available_status,
            isolation_by=admin_user,
            created_by=admin_user,
        )
        return sample

    def _get_preferred_existing_test(self, individual, order=None):
        """Return first matching test by preferred order.

        order: optional list of test type names to prioritize. If None, defaults to [WES, WGS, Targeted Panel].
        """
        preferred = order or ["WES", "WGS", "Targeted Panel"]
        for name in preferred:
            test = individual.get_all_tests().filter(test_type__name=name).order_by("id").first()
            if test:
                return test
        return None

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = options["file_path"]
        admin_username = options.get("admin_username")
        try:
            admin_user = User.objects.get(username=admin_username)
        except User.DoesNotExist:
            self.stdout.write(self.style.ERROR(f"Admin user '{admin_username}' not found"))
            return

        # Ensure required Individual statuses exist
        self._ensure_statuses(admin_user)

        # Load workbook and sheet
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Failed to open workbook: {e}"))
            return

        sheet_name = "yayıniciyedek"
        if sheet_name not in wb.sheetnames:
            self.stdout.write(self.style.ERROR(f"Sheet '{sheet_name}' not found in {file_path}"))
            return
        ws = wb[sheet_name]

        # Header map
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        headers = [h for h in headers if h is not None]

        def row_to_dict(row_vals):
            row = row_vals[: len(headers)]
            return dict(zip(headers, row))

        # Pre-fetch RareBoost IdentifierType if present (optional)
        rareboost_type = IdentifierType.objects.filter(name="RareBoost").first()

        updated_count = 0
        skipped_count = 0

        ct_individual = ContentType.objects.get(app_label="lab", model="individual")

        for values in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in values):
                continue
            data = row_to_dict(values)

            lab_id = data.get("RareBoost ID")
            if not lab_id:
                skipped_count += 1
                continue

            # Locate individual by RareBoost cross-id
            qs = Individual.objects.filter(cross_ids__id_value=str(lab_id))
            if rareboost_type:
                qs = qs.filter(cross_ids__id_type=rareboost_type)
            individual = qs.first()
            if not individual:
                self.stdout.write(f"Individual not found for RareBoost ID: {lab_id}")
                skipped_count += 1
                continue

            changed = False

            # Sex
            sex_val = self._normalize_sex(data.get("Sex"))
            if sex_val and individual.sex != sex_val:
                individual.sex = sex_val
                changed = True

            # is_alive from Status (ex-alive): 1 -> Alive (True), 0 -> Excitus (False)
            alive_field = data.get("Status (ex-alive)")
            alive_bool = self._to_bool(alive_field)
            if alive_bool is not None and individual.is_alive != alive_bool:
                individual.is_alive = alive_bool
                changed = True

            # birth_date only if empty
            if not individual.birth_date:
                dob = self._parse_ddmmyyyy(data.get("Date of Birth"))
                if dob:
                    individual.birth_date = dob
                    changed = True

            # Set Family.is_consanguineous from Consanguinity (1 -> Yes, 0 -> No, UNKNOWN -> skip)
            cons_val = data.get("Consanguinity")
            if individual.family and cons_val not in (None, "UNKNOWN", "Unknown", "unknown"):
                cons_bool = self._to_bool(cons_val)
                if cons_bool is not None and individual.family.is_consanguineous != cons_bool:
                    individual.family.is_consanguineous = cons_bool
                    individual.family.save()

            # NOTE column contains status name; set only if it exists (Individual content type)
            note_status_name = data.get("NOTE")
            if note_status_name:
                note_status_name = str(note_status_name).strip()
                if note_status_name:
                    status_obj = Status.objects.filter(
                        name=note_status_name, content_type=ct_individual
                    ).first()
                    if status_obj and individual.status_id != status_obj.id:
                        individual.status = status_obj
                        changed = True

            if changed:
                individual.save()
                updated_count += 1

            # Create tests from "Previous test" column
            prev_tests_raw = data.get("Previous test")
            if prev_tests_raw:
                text = str(prev_tests_raw)
                # Normalize the quoted token before splitting
                text = text.replace('"Gene Panel, Single Gene"', 'Targeted Panel')
                tokens = [t.strip() for t in text.replace("\n", ",").split(",") if t and t.strip()]
                # Merge residual tokens that belong to Targeted Panel if any
                normalized = []
                skip_next = False
                for idx, tok in enumerate(tokens):
                    if skip_next:
                        skip_next = False
                        continue
                    if tok.lower() == "gene panel" and idx + 1 < len(tokens) and tokens[idx + 1].lower() == "single gene":
                        normalized.append("Targeted Panel")
                        skip_next = True
                    else:
                        normalized.append(tok.strip("\" "))

                ct_test = ContentType.objects.get(app_label="lab", model="test")
                prev_status = Status.objects.get(name="Previous", content_type=ct_test)
                sample_for_tests = self._ensure_sample_for_tests(individual, admin_user)

                for entry in normalized:
                    if not entry:
                        continue
                    tt_name = entry
                    if entry.lower() in {"wes", "wgs", "cma", "karyotype"}:
                        tt_name = entry.upper()
                    if entry.lower() in {"targeted panel", "gene panel single gene"}:
                        tt_name = "Targeted Panel"
                    tt = self._get_or_create_test_type(tt_name, admin_user)
                    # Skip if a test with this type already exists for the individual
                    if individual.get_all_tests().filter(test_type=tt).exists():
                        continue
                    Test.objects.create(
                        test_type=tt,
                        status=prev_status,
                        sample=sample_for_tests,
                        created_by=admin_user,
                    )

            # Handle "RareBoost Reanaliz/WGS/WES/RNA seq"
            rb_raw = data.get("RareBoost Reanaliz/WGS/WES/RNA seq")
            if rb_raw:
                text = str(rb_raw)
                parts = [p.strip() for p in text.replace("\n", ",").split(",") if p and p.strip()]
                # First: if contains Reanalysis or WGS Reanalysis, create an Analysis
                has_reanalysis = any(p.lower() in {"reanalysis", "wgs reanalysis"} for p in parts)
                if has_reanalysis:
                    # Choose preferred order depending on whether WGS Reanalysis is explicitly present
                    if any(p.lower() == "wgs reanalysis" for p in parts):
                        target_test = self._get_preferred_existing_test(
                            individual, order=["WGS", "WES", "Targeted Panel"]
                        )
                    else:
                        target_test = self._get_preferred_existing_test(individual)
                    if not target_test:
                        # If no preferred existing test, try to create WES as default carrier for analysis
                        wes_tt = self._get_or_create_test_type("WES", admin_user)
                        sample_for_tests = self._ensure_sample_for_tests(individual, admin_user)
                        target_test = Test.objects.create(
                            test_type=wes_tt,
                            status=Status.objects.filter(content_type=ContentType.objects.get(app_label="lab", model="test"), name="Completed").first() or Status.objects.filter(content_type=ContentType.objects.get(app_label="lab", model="test")).first(),
                            sample=sample_for_tests,
                            created_by=admin_user,
                        )
                    analysis_type = self._get_or_create_analysis_type("Reanalysis", admin_user)
                    analysis_status = Status.objects.filter(content_type=ContentType.objects.get(app_label="lab", model="analysis"), name="In Progress").first()
                    if not analysis_status:
                        analysis_status = Status.objects.create(
                            name="In Progress",
                            description="Analysis is in progress",
                            color="yellow",
                            created_by=admin_user,
                            content_type=ContentType.objects.get(app_label="lab", model="analysis"),
                            icon="fa-spinner",
                        )
                    Analysis.objects.create(
                        test=target_test,
                        performed_date=datetime.today().date(),
                        performed_by=admin_user,
                        type=analysis_type,
                        status=analysis_status,
                        created_by=admin_user,
                    )
                # Other entries become tests (e.g., WES/WGS/RNA Seq)
                for p in parts:
                    if p.lower() in {"reanalysis", "wgs reanalysis"}:
                        continue
                    name = p
                    if name.lower() in {"wes", "wgs"}:
                        name = name.upper()
                    elif name.lower() in {"rna seq", "rnaseq", "rna-seq"}:
                        name = "RNA Seq"
                    elif name.lower() == "targeted panel":
                        name = "Targeted Panel"
                    tt = self._get_or_create_test_type(name, admin_user)
                    # Create test if doesn't exist
                    if not individual.get_all_tests().filter(test_type=tt).exists():
                        sample_for_tests = self._ensure_sample_for_tests(individual, admin_user)
                        # Default status for newly created tests here: Completed if exists else first Test status
                        test_status = Status.objects.filter(content_type=ContentType.objects.get(app_label="lab", model="test"), name="Completed").first()
                        if not test_status:
                            test_status = Status.objects.filter(content_type=ContentType.objects.get(app_label="lab", model="test")).first()
                        Test.objects.create(
                            test_type=tt,
                            status=test_status,
                            sample=sample_for_tests,
                            created_by=admin_user,
                        )

        self.stdout.write(self.style.SUCCESS(f"Updated {updated_count} individuals; skipped {skipped_count}."))

