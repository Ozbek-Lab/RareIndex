"""Comprehensive data import command for Özbek Lab.

Sheets processed from the master XLSX
---------------------------------------
  OZBEK LAB          – Families, Institutions, Individuals, Samples
  Analiz Takip        – Test + Analysis per row (pipeline linked later by Gennext sheet)
  Variant List        – SNV variants
  Kurumlar            – Institution coordinate / metadata lookup (replaces Gönderen Kurum Harita)
  Sanger Konfirmasyonları – Sanger tests
  WGS_TÜSEB          – WGS tests from TÜSEB
  External            – Individuals without primary/secondary cross-ID
  Katar-Uzun Okuma Hastaları – Long Read WGS (Katar)
  Dubai-Uzun Okuma Hastaları – Long Read WGS (Dubai)
  CP_COHORT           – CP cohort project assignment
  RNA SEQ             – RNA Seq tests
  Gennext Analiz Listesi – Gennext pipelines (links Analiz Takip analyses)
  RarePipe Analiz Listesi – RarePipe pipelines

Optional external inputs
------------------------
  --rarepipe-tsv      – Legacy RarePipe TSV samplesheet
  --yayin-ici         – Path to Yayın_İçi XLSX (sheet GÜNCELyayıniciyedek)
  --forms-dir / --reports-dir – File attachment directories

Processing order
----------------
  0a Load ontologies_data.json fixture (if Ontology table empty)
  0b import_hgnc_data (if Gene table empty)
  1  Setup statuses / IdentifierTypes / ozbek_set_id_priorities
  2  Families + Institutions (OZBEK LAB pass 1)
  3  Individuals + CrossIdentifiers (OZBEK LAB pass 2)
  4  Samples (OZBEK LAB pass 3)
  5  Analiz Takip → Test + Analysis (pipeline=None, linked in step 9)
  6  RarePipe TSV (--rarepipe-tsv)
  7  Parent links
  8  Sanger Konfirmasyonları
  9  WGS_TÜSEB
 10  External
 11  Katar / Dubai long-read sheets
 12  CP_COHORT
 13  RNA SEQ
 14  Gennext Analiz Listesi (creates Pipelines and links to analyses from step 5)
 15  RarePipe Analiz Listesi
 16  Variant List
 17  link_imported_genes
 18  File attachments
 19  Yayın_İçi (--yayin-ici)

REMINDERS (ask after implementation):
  • CNV / SV / Repeat variant format in Variant List (Q5)
  • RarePipe Analiz Listesi: confirm whether Matching Sample ID / ID should also be preserved as notes
  • Yayın_İçi Variant column format (Q12)
"""

import csv
import json
import os
import re
from pathlib import Path

import openpyxl
from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.core.files import File
from django.core.management import call_command
from django.core.management.base import BaseCommand, CommandError

from lab.models import (
    Analysis,
    AnalysisReport,
    AnalysisRequestForm,
    CrossIdentifier,
    Contact,
    Family,
    IdentifierType,
    Individual,
    Institution,
    Pipeline,
    Project,
    Sample,
    PlotTemplate,
    Status,
    Task,
    Test,
)
from ontologies.models import Ontology
from variant.models import Classification, CNV, Gene, SNV, SV, Variant, delins

from lab.management.commands._import_helpers import (
    build_id_map,
    find_individual_by_import_identifier,
    find_individual_by_rareboost_id,
    get_family_id,
    get_hpo_terms,
    get_initials,
    get_or_create_analysis_type,
    get_or_create_pipeline_type,
    get_or_create_sample_type,
    get_or_create_status,
    get_or_create_status_group,
    get_or_create_test_type,
    get_or_create_contact,
    get_or_create_contact_for_user,
    get_or_create_user,
    identifier_type_example_for_name,
    map_classification,
    map_inheritance,
    normalize_id,
    normalize_sex,
    parse_and_add_notes,
    parse_date,
    parse_date_from_filename,
    parse_variant_string,
    to_bool,
)

User = get_user_model()


def normalize_consanguinity_value(value):
    """Map imported consanguinity values to a nullable boolean."""
    bool_value = to_bool(value)
    if bool_value is not None:
        return bool_value

    text = str(value or "").strip().lower()
    if not text:
        return None
    if text in {"mild", "mildly", "mildly consanguineous", "hafif", "hafif akraba", "hafif akrabalık"}:
        return None
    if text in {"unknown", "bilinmiyor", "belirsiz", "na", "n/a", "-"}:
        return None
    return None


# ---------------------------------------------------------------------------
# Zygosity map (spreadsheet labels → model choice keys)
# ---------------------------------------------------------------------------
ZYGOSITY_MAP = {
    "het.": "het",
    "heterozigot": "het",
    "hom.": "hom",
    "homozigot": "hom",
    "hemizigot": "hemi",
    "heteroplazmi": "hetpl",
    "homoplazmi": "homoplasmy",
    "na": "unknown",
    "n/a": "unknown",
}

YAYIN_ZYGOSITY_MAP = {
    "het": "het",
    "het.": "het",
    "heterozygous": "het",
    "heterozigot": "het",
    "hom": "hom",
    "hom.": "hom",
    "homozygous": "hom",
    "homozigot": "hom",
    "hem": "hemi",
    "hemi": "hemi",
    "hemizigot": "hemi",
    "heteroplazmi": "hetpl",
    "hetpl": "hetpl",
    "homoplazmi": "homoplasmy",
    "comphat": "het",
    "comphet": "het",
    "mono - de novo": "unknown",
    "mono-de novo": "unknown",
    "mono - maternal": "unknown",
    "mono-maternal": "unknown",
    "mono - paternal": "unknown",
    "mono-paternal": "unknown",
    "mono - unknown": "unknown",
    "mono-unknown": "unknown",
    "negative": "unknown",
    "unknown": "unknown",
    "na": "unknown",
    "n/a": "unknown",
}

REQUIRED_PLOT_TEMPLATE_SPECS = {
    "sample-distribution-sunburst": {
        "default_col_span": 1,
        "show_download_menu": False,
    },
    "custom-sunburst": {
        "default_col_span": 2,
        "show_download_menu": False,
    },
    "analysis-status-bar": {
        "default_col_span": 1,
        "show_download_menu": False,
    },
    "hpo-term-network": {
        "default_col_span": 2,
        "show_download_menu": False,
    },
}

REQUIRED_PLOT_TEMPLATE_SLUGS = set(REQUIRED_PLOT_TEMPLATE_SPECS)

ANALYSIS_IMPORT_STATUS_NAMES = {
    "reported": "Reported",
    "waiting analysis": "Waiting Analysis",
    "waiting_analysis": "Waiting Analysis",
    "planned": "Planned",
    "plan": "Planned",
    "performed": "Performed",
    "analyzed": "Analyzed",
    "analysed": "Analyzed",
    "waiting data arrival": "Waiting Data Arrival",
    "waiting_data_arrival": "Waiting Data Arrival",
}


def _parse_report_text_field_reference_markdown(md_text: str) -> dict[str, dict[str, str]]:
    """
    Parse `report_text_field_reference.md` into:
      { "WES": { "positive_report_template": "...", ... }, ... }
    """
    lines = md_text.splitlines()
    parsed: dict[str, dict[str, str]] = {}
    current_test: str | None = None

    key_re = re.compile(r'^`(?P<key>[a-zA-Z0-9_]+)`\s*$')
    section_re = re.compile(r'^##\s+(?P<name>.+)\s*$')
    inline_value_re = re.compile(r'^`(?P<val>[^`]*)`\s*$')

    i = 0
    while i < len(lines):
        line = lines[i].rstrip("\n")

        sec_m = section_re.match(line)
        if sec_m:
            current_test = sec_m.group("name").strip()
            parsed.setdefault(current_test, {})
            i += 1
            continue

        if current_test is None:
            i += 1
            continue

        key_m = key_re.match(line.strip())
        if not key_m:
            i += 1
            continue

        field_name = key_m.group("key")

        j = i + 1
        while j < len(lines) and lines[j].strip() == "":
            j += 1
        if j >= len(lines):
            raise RuntimeError(f"Unexpected EOF while reading value for {current_test}.{field_name}")

        val_line = lines[j].strip()
        if val_line.startswith("```"):
            j += 1
            content: list[str] = []
            while j < len(lines):
                if lines[j].strip().startswith("```"):
                    break
                content.append(lines[j])
                j += 1
            else:
                raise RuntimeError(f"Missing closing fence for {current_test}.{field_name}")

            value = "\n".join(content).strip("\n")
            parsed[current_test][field_name] = value
            i = j + 1
            continue

        in_m = inline_value_re.match(val_line)
        if in_m:
            parsed[current_test][field_name] = in_m.group("val")
            i = j + 1
            continue

        parsed[current_test][field_name] = val_line
        i = j + 1

    return parsed


def _normalize_testtype_report_payload(payload: dict[str, str]) -> dict[str, str]:
    normalized = dict(payload)

    fallback_pairs = (
        ("default_method_text", ("positive_method_text", "negative_method_text")),
        ("default_filtering_text", ("positive_filtering_text", "negative_filtering_text")),
        ("default_limitations_text", ("positive_limitations_text", "negative_limitations_text")),
    )
    for target, candidates in fallback_pairs:
        if normalized.get(target):
            continue
        for candidate in candidates:
            candidate_value = normalized.get(candidate)
            if candidate_value:
                normalized[target] = candidate_value
                break

    return normalized


def _map_zygosity_strict(value, warn_fn=None):
    """Return model key for *value*, or None and call warn_fn if unrecognised."""
    if not value:
        if warn_fn:
            warn_fn(f"  Zygosity is empty — skipping variant")
        return None
    normalized = re.sub(r"\s+", " ", str(value).strip().lower())
    mapped = ZYGOSITY_MAP.get(normalized)
    if mapped is None and warn_fn:
        warn_fn(f"  Unrecognised zygosity {value!r} — skipping variant")
    return mapped


def _compact_variant_coord(value) -> str:
    """Strip punctuation commonly used as thousands separators from coordinate strings."""
    if value is None:
        return ""
    return re.sub(r"[,\s.]", "", str(value).strip())


def _normalize_variant_chromosome(value) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text.lower().startswith("chr"):
        return text if text.startswith("chr") else f"chr{text[3:]}"
    m = re.match(r"^(?P<num>\d+|x|y|m)", text, re.I)
    if m:
        suffix = m.group("num")
        return f"chr{suffix.upper() if len(suffix) == 1 else suffix}"
    return text


def _normalize_yayin_zygosity(value) -> str:
    if value is None:
        return "unknown"
    text = re.sub(r"\s+", " ", str(value).strip().lower())
    text = re.sub(r"\s+", " ", text)
    return YAYIN_ZYGOSITY_MAP.get(text, YAYIN_ZYGOSITY_MAP.get(text.replace(" ", "-"), "unknown"))


def _split_yayin_variant_text(value) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text or text == "-":
        return []
    return [line.strip() for line in text.splitlines() if line.strip() and line.strip() != "-"]


def _split_csv_values(value) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def _normalize_contact_value(value) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def _build_clinician_assignments(
    clinician_field,
    first_contact_field,
    second_contact_field,
) -> list[tuple[str, list[str]]]:
    names = [name.strip() for name in _split_csv_values(clinician_field)]
    names = [name for name in names if name]
    if not names:
        return []

    contacts = []
    for raw_field in (first_contact_field, second_contact_field):
        for token in _split_csv_values(raw_field):
            normalized = _normalize_contact_value(token)
            if normalized:
                contacts.append(normalized)

    assignments = [[] for _ in names]
    for idx, contact in enumerate(contacts):
        assignments[min(idx, len(names) - 1)].append(contact)

    return list(zip(names, assignments))


def _extract_variant_records(token: str) -> list[dict]:
    """
    Best-effort parser for Yayın_İçi variant text.

    Returns a list of dicts describing importable variants. Unparseable tokens
    are returned as an empty list so callers can preserve them as notes.
    """
    text = str(token or "").strip()
    if not text or text == "-":
        return []

    variant_text, sep, extra_text = text.partition(";")
    variant_text = variant_text.strip()
    extra_text = extra_text.strip() if sep else ""

    # Gene-level HGVS-like strings without genomic coordinates are preserved as notes.
    if "chr" not in variant_text.lower() and not re.search(r"\(\s*[\d,._]+\s*[_-]\s*[\d,._]+\s*\)\s*x\d+", variant_text, re.I):
        return []

    # Explicit "chr5-12345 A>G" or "chr5:12345 A>G"
    snv_match = re.search(
        r"(?P<chrom>chr[\w]+)[\s:,-]+(?P<start>[\d.,]+)\s*(?P<ref>[ACGT]+)>(?P<alt>[ACGT]+)",
        variant_text,
        re.I,
    )
    if snv_match:
        ref = snv_match.group("ref").upper()
        alt = snv_match.group("alt").upper()
        record = {
            "chromosome": _normalize_variant_chromosome(snv_match.group("chrom")),
            "start": int(_compact_variant_coord(snv_match.group("start"))),
            "reference": ref,
            "alternate": alt,
            "kind": "snv" if len(ref) == len(alt) == 1 else "delins",
            "end": int(_compact_variant_coord(snv_match.group("start"))),
        }
        if extra_text:
            record["note"] = extra_text
        record["source_text"] = text
        return [record]

    # Structural/CNV style variants with a coordinate range.
    coord_range_match = re.search(
        r"(?P<chrom>chr[\w]+|[0-9XYM]+[pq][^(\s]*)[:\-](?P<start>[\d.,]+)[\-_](?P<end>[\d.,]+)",
        variant_text,
        re.I,
    )
    if coord_range_match:
        copy_match = re.search(r"x(?P<copy>\d+)", variant_text, re.I)
        has_dup = re.search(r"\bdup(lication)?\b|\bgain\b|\bamplification\b", variant_text, re.I)
        has_del = re.search(r"\bdel(etion)?\b|\bdelesyon\b", variant_text, re.I)
        if not (copy_match or has_dup or has_del):
            return []

        chrom = _normalize_variant_chromosome(coord_range_match.group("chrom"))
        start = int(_compact_variant_coord(coord_range_match.group("start")))
        end = int(_compact_variant_coord(coord_range_match.group("end")))
        if copy_match:
            copy = int(copy_match.group("copy"))
            return [{
                "chromosome": chrom,
                "start": min(start, end),
                "end": max(start, end),
                "kind": "cnv",
                "cnv_type": "loss" if copy == 0 else "gain",
                "copy_number": copy,
                "source_text": text,
                **({"note": extra_text} if extra_text else {}),
            }]
        variant_type = "sv"
        if has_dup:
            variant_type = "cnv"

        record = {
            "chromosome": chrom,
            "start": min(start, end),
            "end": max(start, end),
            "kind": variant_type,
            "source_text": text,
        }
        if extra_text:
            record["note"] = extra_text
        if variant_type == "cnv":
            record["cnv_type"] = "gain"
            record["copy_number"] = None
        elif variant_type == "sv":
            record["sv_type"] = "deletion" if has_del else "structural_variant"
        return [record]

    # Cytoband / deletion notation with coordinates in parentheses:
    #   LAMA2 seq[GRCh38] 6q22.33(129,047,209_129,083,546)x4 Duplikasyon
    band_match = re.search(
        r"(?P<chrom>[0-9XYM]+[pq][^(\s]*)\((?P<start>[\d,._]+)\s*[_-]\s*(?P<end>[\d,._]+)\)\s*x(?P<copy>\d+)",
        variant_text,
        re.I,
    )
    if band_match:
        chrom = _normalize_variant_chromosome(band_match.group("chrom"))
        start = int(_compact_variant_coord(band_match.group("start")))
        end = int(_compact_variant_coord(band_match.group("end")))
        copy = int(band_match.group("copy"))
        return [{
            "chromosome": chrom,
            "start": min(start, end),
            "end": max(start, end),
            "kind": "cnv",
            "cnv_type": "loss" if copy == 0 else "gain",
            "copy_number": copy,
            "source_text": text,
            **({"note": extra_text} if extra_text else {}),
        }]

    return []


def _variant_text_summary(token: str) -> str:
    token = " ".join(str(token or "").split())
    return token.strip()


class Command(BaseCommand):
    help = "Comprehensive Özbek Lab data import (XLSX + optional external files)"

    # ------------------------------------------------------------------
    # Arguments
    # ------------------------------------------------------------------

    def add_arguments(self, parser):
        parser.add_argument("xlsx_file", type=str, help="Path to the master XLSX file")
        parser.add_argument("--admin-username", required=True,
                            help="Username for created_by / performed_by fallback")
        parser.add_argument("--rarepipe-tsv", dest="rarepipe_tsv",
                            help="Path to legacy RarePipe TSV samplesheet")
        parser.add_argument("--yayin-ici", dest="yayin_ici",
                            help="Path to Yayın_İçi XLSX (GÜNCELyayıniciyedek sheet)")
        parser.add_argument("--forms-dir", dest="forms_dir",
                            help="Directory of analysis request form files")
        parser.add_argument("--reports-dir", dest="reports_dir",
                            help="Directory of analysis report files")
        parser.add_argument("--skip-hgnc", dest="skip_hgnc", action="store_true",
                            help="Skip HGNC gene data download check")
        parser.add_argument("--dry-run", dest="dry_run", action="store_true",
                            help="Validate without writing to the database")

    def _resolve_admin_user(self, admin_username: str):
        """Get the admin user, creating or promoting it when needed."""
        admin_user = User.objects.filter(username=admin_username).first()
        if admin_user:
            needs_staff = not admin_user.is_staff
            needs_superuser = not admin_user.is_superuser
            if needs_staff or needs_superuser:
                if self.dry_run:
                    self.stdout.write(self.style.WARNING(
                        f"Admin user {admin_username!r} exists but is not staff/superuser; "
                        "a real run would promote it."
                    ))
                else:
                    admin_user.is_staff = True
                    admin_user.is_superuser = True
                    admin_user.save(update_fields=["is_staff", "is_superuser"])
                    self.stdout.write(self.style.WARNING(
                        f"Admin user {admin_username!r} was promoted to staff/superuser."
                    ))
            return admin_user

        if self.dry_run:
            raise CommandError(
                f"Admin user {admin_username!r} not found and cannot be created in dry-run mode"
            )

        admin_user = User.objects.create_superuser(
            username=admin_username,
            password=admin_username,
        )
        self.stdout.write(self.style.WARNING(
            f"Admin user {admin_username!r} did not exist and was created as a superuser."
        ))
        return admin_user

    def _apply_contact_details(self, contact, contact_values) -> None:
        if not contact_values:
            return

        existing_emails = list(contact.emails or [])
        existing_phones = list(contact.phones or [])
        updated_emails = list(existing_emails)
        updated_phones = list(existing_phones)

        for value in contact_values:
            normalized_value = _normalize_contact_value(value)
            if not normalized_value:
                continue
            if "@" in normalized_value:
                if normalized_value not in updated_emails:
                    updated_emails.append(normalized_value)
            elif normalized_value not in updated_phones:
                updated_phones.append(normalized_value)

        changed_fields = []
        if updated_emails != existing_emails:
            contact.emails = updated_emails
            changed_fields.append("emails")
        if updated_phones != existing_phones:
            contact.phones = updated_phones
            changed_fields.append("phones")
        if changed_fields:
            contact.save(update_fields=changed_fields)

        linked_user = getattr(contact, "user", None)
        if linked_user and not linked_user.email:
            email_to_set = next((email for email in updated_emails if "@" in email), None)
            if email_to_set:
                linked_user.email = email_to_set
                linked_user.save(update_fields=["email"])

    def _link_physician_to_individual_and_institutions(self, individual, physician, institutions) -> None:
        """Link an imported clinician to the individual and the individual's source institutions."""
        individual.physicians.add(physician)
        for institution in institutions or []:
            if institution:
                institution.staff.add(physician)

    def _build_clinician_assignments(self, clinician_field, first_contact_field, second_contact_field):
        assignments = _build_clinician_assignments(
            clinician_field,
            first_contact_field,
            second_contact_field,
        )
        self._report_clinician_edge_cases(
            clinician_field,
            first_contact_field,
            second_contact_field,
            assignments,
        )
        return assignments

    def _report_clinician_edge_cases(
        self,
        clinician_field,
        first_contact_field,
        second_contact_field,
        assignments,
    ) -> None:
        clinician_names = [name.strip() for name in _split_csv_values(clinician_field) if name.strip()]
        contact_values = []
        for raw_field in (first_contact_field, second_contact_field):
            for token in _split_csv_values(raw_field):
                normalized = _normalize_contact_value(token)
                if normalized:
                    contact_values.append(normalized)

        if not clinician_names or not contact_values:
            return

        if len(contact_values) < len(clinician_names):
            message = (
                "Clinician contact edge case: "
                f"{len(clinician_names)} clinician(s) but {len(contact_values)} contact value(s); "
                "later clinicians will have no contact info."
            )
        elif len(contact_values) > len(clinician_names):
            message = (
                "Clinician contact edge case: "
                f"{len(clinician_names)} clinician(s) but {len(contact_values)} contact value(s); "
                "extra values were attached to the last clinician."
            )
        else:
            return

        self.stdout.write(f"INFO: {message}")
        self._record_issue(
            step="clinician",
            sheet="Klinisyen",
            severity="info",
            reason=message,
            row={
                "Klinisyen": clinician_field,
                "İletişim Bilgileri - Mail/telefon?": first_contact_field,
                "İletişim Bilgileri - Telefon/mail?": second_contact_field,
            },
            context={"assignments": assignments},
        )

    # ------------------------------------------------------------------
    # Main entry point
    # ------------------------------------------------------------------

    def handle(self, *args, **options):
        self.dry_run: bool = options["dry_run"]
        if self.dry_run:
            self.stdout.write(self.style.WARNING("-- DRY RUN: nothing will be saved --"))

        admin_username = options["admin_username"]
        self.admin_user = self._resolve_admin_user(admin_username)

        file_path = options["xlsx_file"]
        if not os.path.exists(file_path):
            raise CommandError(f"File not found: {file_path}")

        self._init_issue_log(file_path)

        # Instance-level state shared between steps
        self.analysis_map: dict = {}       # (lab_id, tt_name) → Analysis

        # Lazily loaded on-demand by _backfill_testtype_report_fields().
        self._report_text_reference_cache = None

        # Step 0a — ontologies
        self._step0a_ensure_ontologies()

        # Step 0b — HGNC genes
        self._step0b_ensure_hgnc(options["skip_hgnc"])

        # Step 1
        self.statuses, self.id_types = self._step1_setup()

        # Load workbook
        self.stdout.write(f"Loading workbook: {file_path}")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        kurumlar_map = self._load_kurumlar_map(wb)
        rows = self._load_ozbek_lab_rows(wb)

        # Steps 2–4 — OZBEK LAB
        families, institutions, unknown_inst = self._step2_families_institutions(
            rows, kurumlar_map)
        self._step3_individuals(rows, families, institutions, unknown_inst)
        self._step4_samples(rows)

        # Step 5 — Analiz Takip
        self._step5_analiz_takip(wb)

        # Step 6 — RarePipe TSV (external)
        if options.get("rarepipe_tsv"):
            self._step6_rarepipe(options["rarepipe_tsv"])

        # Step 7 — Parent links
        self._step7_parent_links(families)

        # Steps 8–15 — extra sheets
        self._step_sanger(wb)
        self._step_wgs_tuseb(wb)
        self._step_external(wb, kurumlar_map)
        self._step_long_read(wb, "Katar-Uzun Okuma Hastaları", "Katar",
                             "Qatar - Long Read WGS Project")
        self._step_long_read(wb, "Dubai-Uzun Okuma Hastaları", "Dubai",
                             "Dubai - Long Read WGS Project")
        # self._step_cp_cohort(wb)
        self._step_rna_seq(wb)
        self._step_gennext_analiz(wb)       # links Analiz Takip analyses to pipelines
        self._step_rarepipe_analiz(wb)      # ⚠ skipped — no date column yet

        # Step 16 — Variant List
        self._step_variants(wb)

        # Step 17 — link_imported_genes
        if not self.dry_run:
            self.stdout.write("Step 17: Linking genes via annotations…")
            call_command("link_imported_genes")

        # Step 18 — plot templates
        self._step18_ensure_plot_templates()

        # Step 19 — file attachments
        if options.get("forms_dir") or options.get("reports_dir"):
            self._step_file_attachments(options.get("forms_dir"), options.get("reports_dir"))

        # Step 20 — Yayın_İçi
        if options.get("yayin_ici"):
            self._step_yayin_ici(options["yayin_ici"])

        self._write_issue_log()
        self.stdout.write(self.style.SUCCESS("Import completed successfully."))

    def _init_issue_log(self, xlsx_file: str) -> None:
        base_dir = Path(xlsx_file).resolve().parent
        self.issue_log_path = base_dir / "import_all_issues.tsv"
        self.info_log_path = base_dir / "import_all_info.tsv"
        self.error_log_path = base_dir / "import_all_errors.tsv"
        self.issue_records: list[dict[str, str]] = []

    def _record_issue(
        self,
        *,
        step: str,
        reason: str,
        sheet: str = "",
        severity: str = "warning",
        lab_id=None,
        row: dict | None = None,
        context: dict | None = None,
    ) -> None:
        def _json_safe(value):
            if isinstance(value, dict):
                return {
                    "" if key is None else str(key): _json_safe(val)
                    for key, val in value.items()
                }
            if isinstance(value, (list, tuple, set)):
                return [_json_safe(item) for item in value]
            return value

        payload = {
            "step": step,
            "sheet": sheet,
            "severity": severity,
            "reason": reason,
            "lab_id": "" if lab_id is None else str(lab_id),
            "context": json.dumps(_json_safe(context or {}), ensure_ascii=False, sort_keys=True, default=str),
            "row_data": json.dumps(_json_safe(row or {}), ensure_ascii=False, sort_keys=True, default=str),
        }
        self.issue_records.append(payload)

    def _write_issue_log(self) -> None:
        records = list(getattr(self, "issue_records", []) or [])

        def _sanitize_shareable_record(record: dict[str, str]) -> dict[str, str]:
            return {
                "step": record.get("step", ""),
                "sheet": record.get("sheet", ""),
                "severity": record.get("severity", ""),
                "reason": record.get("reason", ""),
                "lab_id": record.get("lab_id", ""),
            }

        non_info_records = [record for record in records if record.get("severity") != "info"]
        info_records = [record for record in records if record.get("severity") == "info"]
        error_records = [record for record in records if record.get("severity") == "error"]

        def _write_tsv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
            with path.open("w", encoding="utf-8", newline="") as fh:
                writer = csv.DictWriter(fh, fieldnames=fieldnames, delimiter="\t")
                writer.writeheader()
                writer.writerows(rows)

        _write_tsv(
            self.issue_log_path,
            non_info_records,
            ["step", "sheet", "severity", "reason", "lab_id", "context", "row_data"],
        )
        _write_tsv(
            self.info_log_path,
            [_sanitize_shareable_record(record) for record in info_records],
            ["step", "sheet", "severity", "reason", "lab_id"],
        )
        _write_tsv(
            self.error_log_path,
            [_sanitize_shareable_record(record) for record in error_records],
            ["step", "sheet", "severity", "reason", "lab_id"],
        )

        if records:
            self.stdout.write(self.style.WARNING(
                f"Issue log: wrote {len(non_info_records)} non-info entries to {self.issue_log_path}"))
            self.stdout.write(self.style.WARNING(
                f"Issue log: wrote {len(info_records)} info entries to {self.info_log_path}"))
            self.stdout.write(self.style.WARNING(
                f"Issue log: wrote {len(error_records)} error entries to {self.error_log_path}"))
        else:
            self.stdout.write("Issue log: no skipped or failed rows recorded.")

    def _fit_uploaded_filename(self, field_file, source_name: str) -> str:
        """
        Shorten the basename if needed so the generated storage path fits the
        field's max_length. Keeps the extension intact.
        """
        field = field_file.field
        max_length = field.max_length
        candidate_name = Path(source_name).name
        generated_name = field.generate_filename(field_file.instance, candidate_name)

        if not max_length or len(generated_name) <= max_length:
            return candidate_name

        suffix = Path(candidate_name).suffix
        stem = Path(candidate_name).stem
        min_stem_len = 1
        allowed_stem_len = len(stem)

        while allowed_stem_len >= min_stem_len:
            trimmed_name = f"{stem[:allowed_stem_len]}{suffix}"
            generated_name = field.generate_filename(field_file.instance, trimmed_name)
            if len(generated_name) <= max_length:
                return trimmed_name
            allowed_stem_len -= 1

        if suffix:
            trimmed_name = suffix[-max_length:]
            generated_name = field.generate_filename(field_file.instance, trimmed_name)
            if len(generated_name) <= max_length:
                return trimmed_name

        raise ValueError(
            f"Could not fit filename {source_name!r} into max_length={max_length} "
            f"for upload path {field.upload_to!r}"
        )

    def _load_report_text_reference(self) -> dict[str, dict[str, str]]:
        if self._report_text_reference_cache is not None:
            return self._report_text_reference_cache

        repo_root = Path(__file__).resolve().parents[3]
        md_path = repo_root / "report_text_field_reference.md"
        if not md_path.exists():
            self._report_text_reference_cache = {}
            return self._report_text_reference_cache

        self._report_text_reference_cache = _parse_report_text_field_reference_markdown(
            md_path.read_text(encoding="utf-8")
        )
        return self._report_text_reference_cache

    def _backfill_testtype_report_fields(self, test_type) -> None:
        """
        Populate empty TestType report fields from the markdown reference.
        Uses bulk UPDATE to avoid unnecessary model save side-effects.
        """
        if self.dry_run:
            return
        if not test_type or not getattr(test_type, "pk", None):
            return
        if not getattr(test_type, "name", None):
            return

        parsed = self._load_report_text_reference()
        if not parsed:
            return

        lower_to_canonical = {k.strip().lower(): k for k in parsed.keys()}
        canonical_name = lower_to_canonical.get(test_type.name.strip().lower())
        if not canonical_name:
            return

        payload = _normalize_testtype_report_payload(parsed.get(canonical_name) or {})
        model_fields = {f.name for f in type(test_type)._meta.fields}

        updates = {}
        for field_name, value in payload.items():
            if field_name not in model_fields:
                continue
            current = getattr(test_type, field_name, "")
            if current is None or (isinstance(current, str) and current.strip() == ""):
                updates[field_name] = value

        if updates:
            type(test_type).objects.filter(pk=test_type.pk).update(**updates)

    def _step18_ensure_plot_templates(self) -> None:
        """
        Ensure the published Marimo-backed plot templates exist so /gallery/
        can show cards and link to the run/editor servers.
        """
        existing_templates = {
            slug: {
                "default_col_span": default_col_span,
                "show_download_menu": show_download_menu,
            }
            for slug, default_col_span, show_download_menu in PlotTemplate.objects.filter(
                slug__in=REQUIRED_PLOT_TEMPLATE_SLUGS,
                is_published=True,
            ).values_list("slug", "default_col_span", "show_download_menu")
        }
        missing_slugs = sorted(REQUIRED_PLOT_TEMPLATE_SLUGS - set(existing_templates))
        mismatched_specs = sorted(
            slug
            for slug, spec in REQUIRED_PLOT_TEMPLATE_SPECS.items()
            if any(existing_templates.get(slug, {}).get(field) != value for field, value in spec.items())
        )
        if not missing_slugs and not mismatched_specs:
            self.stdout.write("Step 18: Plot templates already seeded.")
            return

        self.stdout.write(
            self.style.WARNING(
                f"Step 18: Plot templates need seeding (missing={missing_slugs}, spec_mismatch={mismatched_specs}) — seeding defaults…"
            )
        )
        if self.dry_run:
            return

        call_command("seed_plot_templates")

    # ==================================================================
    # Step 0a — Ontologies
    # ==================================================================

    def _step0a_ensure_ontologies(self) -> None:
        """Load the bundled ontologies_data.json fixture when the Ontology
        table is empty so that HPO terms are available for the rest of the
        import (individuals, Yayın_İçi HPO attachments, etc.)."""
        if Ontology.objects.exists():
            self.stdout.write("Step 0a: Ontologies already loaded.")
            return
        self.stdout.write(self.style.WARNING(
            "Step 0a: Ontology table empty — loading ontologies_data.json fixture…"))
        if not self.dry_run:
            call_command("loaddata", "ontologies_data.json")
            loaded = Ontology.objects.count()
            if loaded:
                self.stdout.write(self.style.SUCCESS(
                    f"Step 0a: Loaded {loaded} ontologies."))
            else:
                self.stdout.write(self.style.WARNING(
                    "Step 0a: Fixture loaded but no Ontology objects found. "
                    "Ensure ontologies_data.json is in a fixtures/ directory."))

    # ==================================================================
    # Step 0b — HGNC
    # ==================================================================

    def _step0b_ensure_hgnc(self, skip_hgnc: bool) -> None:
        if skip_hgnc:
            self.stdout.write("Step 0b: --skip-hgnc set.")
            return
        if Gene.objects.exists():
            self.stdout.write("Step 0b: Gene table populated — skipping import_hgnc_data.")
            return
        self.stdout.write(self.style.WARNING(
            "Step 0b: Gene table empty — running import_hgnc_data…"))
        if not self.dry_run:
            call_command("import_hgnc_data")

    # ==================================================================
    # Step 1 — Setup
    # ==================================================================

    def _step1_setup(self) -> tuple:
        self.stdout.write("Step 1: Setting up statuses and identifier types…")
        ct = self._content_types()

        def ct_for(key):
            return ct.get(key)

        def group_for(ct_key, group_name):
            return get_or_create_status_group(group_name, ct_for(ct_key)) if group_name else None

        def s(
            name,
            desc,
            color,
            ct_key,
            icon,
            short_name="",
            group_name=None,
            connected_class_keys=(),
        ):
            connected_classes = [ct_for(key) for key in connected_class_keys if ct_for(key)]
            return get_or_create_status(
                name,
                desc,
                color,
                self.admin_user,
                ct_for(ct_key),
                icon,
                short_name=short_name,
                group=group_for(ct_key, group_name),
                connected_classes=connected_classes or None,
            )

        statuses = {
            "individual": {
                "active":       s("Active",       "Imported / manually active",    "green",  "individual", "fa-user-check", short_name="Act", group_name="Activity"),
                "inactive":     s("Inactive",     "Manually inactive",             "gray",   "individual", "fa-user-slash", short_name="Ina", group_name="Activity"),
                "affected":     s("Affected",     "HPO present",                  "red",    "individual", "fa-disease", short_name="Aff", group_name="Affectedness"),
                "healthy":      s("Healthy",      "No HPO currently set",         "green",  "individual", "fa-heart", short_name="Hea", group_name="Affectedness"),
                "unsolved":     s("Unsolved",     "Entry unsolved",               "brown",  "individual", "fa-circle-xmark", short_name="Uns", group_name="Solved"),
                "solved":       s("Solved",       "Entry solved",                 "green",  "individual", "fa-circle-check", short_name="Sol", group_name="Solved"),
                "unsure_import": s("Unsure Import", "Imported from uncertain source", "orange", "individual", "fa-circle-question"),
            },
            "sample": {
                "planned":          s("Planned",                   "Queued for processing",     "yellow", "sample", "fa-calendar", short_name="Plan", group_name="Process"),
                "received":         s("Recieved - In lab process", "Received and processing",   "orange", "sample", "fa-vials", short_name="Rec", group_name="Process"),
                "isolated":         s("Isolated",                  "Isolation complete",        "green",  "sample", "fa-circle-check", short_name="Iso", group_name="Process"),
                "available":        s("Available",                "Available for use",         "green",  "sample", "fa-circle-check", short_name="Ava", group_name="Availability"),
                "not_available":    s("Not Available",            "Not available",             "red",    "sample", "fa-ban", short_name="N/A", group_name="Availability", connected_class_keys=("individual",)),
                "unsure_import":    s("Unsure Import",            "Import fallback",           "orange", "sample", "fa-circle-question"),
            },
            "test": {
                "planned":      s("Planned",                          "Waiting to start",          "yellow", "test", "fa-flask", short_name="Plan", group_name="Process", connected_class_keys=("individual",)),
                "waiting":      s("Waiting Data/Bioinformatic process","Awaiting data / bioinfo",   "orange", "test", "fa-spinner", short_name="Wait", group_name="Process"),
                "completed":    s("Data Delivered / Completed",       "Test completed",            "green",  "test", "fa-circle-check", short_name="Comp", group_name="Process"),
                "previous":     s("Previous",                         "Historical",                "grey",   "test", "fa-clock-rotate-left", short_name="Prev", group_name="Previous"),
                "unsure_import": s("Unsure Import",                   "Import fallback",           "orange", "test", "fa-circle-question"),
            },
            "pipeline": {
                "planned":       s("Planned",                           "Queued",                     "yellow", "pipeline", "fa-diagram-project", short_name="Plan", group_name="Process"),
                "waiting":       s("Waiting Data/Bioinformatic process","Awaiting data / bioinfo",    "orange", "pipeline", "fa-spinner", short_name="Wait", group_name="Process"),
                "completed":     s("Bioinformatic process completed",   "Pipeline completed",         "green",  "pipeline", "fa-circle-check", short_name="Comp", group_name="Process"),
                "unsure_import": s("Unsure Import",                    "Import fallback",            "orange", "pipeline", "fa-circle-question"),
            },
            "analysis": {
                "planned":       s("Planned",                  "Queued",                     "yellow", "analysis", "fa-calendar", short_name="PLan", group_name="Process"),
                "waiting":       s("Waiting Confirmation",    "Awaiting review",            "orange", "analysis", "fa-spinner", short_name="Conf", group_name="Process"),
                "completed":     s("Completed",               "Analysis completed",         "blue",   "analysis", "fa-circle-check", short_name="Comp", group_name="Process", connected_class_keys=("individual",)),
                "reported":      s("Reported",                "Reported to clinician",      "green",  "analysis", "fa-file-circle-check", short_name="Rep", group_name="Process"),
                "initial":       s("Initial Analysis",        "Initial analysis",           "orange", "analysis", "fa-seedling", short_name="Int", group_name="Occasion"),
                "reanalysis":    s("Reanalysis",              "Reanalysis",                 "yellow", "analysis", "fa-rotate", short_name="Rea", group_name="Occasion"),
                "unsure_import": s("Unsure Import",            "Import fallback",            "orange", "analysis", "fa-circle-question"),
            },
            "analysisreport": {
                "negative":      s("Negative",                "Negative result",           "red",    "analysisreport", "fa-circle-xmark", short_name="Neg", group_name="Result", connected_class_keys=("individual",)),
                "positive":      s("Positive",                "Positive result",           "green",  "analysisreport", "fa-circle-check", short_name="Pos", group_name="Result", connected_class_keys=("individual",)),
                "delivered":     s("Delivered to Clinician",  "Delivered to clinician",    "green",  "analysisreport", "fa-envelope-open-text", short_name="Del", group_name="Informed"),
                "unsure_import": s("Unsure Import",           "Import fallback",           "orange", "analysisreport", "fa-circle-question"),
            },
            "project": {
                "in_planning": s("In Planning", "Planning", "blue",   "project", "fa-compass-drafting", short_name="Plan", group_name="Process"),
                "in_progress": s("In Progress", "Active",   "yellow", "project", "fa-diagram-project", short_name="Prog", group_name="Process"),
                "on_hold":     s("On Hold",     "Paused",   "orange", "project", "fa-pause", short_name="Hold", group_name="Process"),
                "completed":   s("Completed",   "Done",     "green",  "project", "fa-flag-checkered", short_name="Comp", group_name="Process"),
                "cancelled":   s("Cancelled",   "Cancelled","grey",   "project", "fa-ban", short_name="Canc", group_name="Process"),
            },
            "task": {
                "assigned":  s("Assigned",  "Assigned", "yellow", "task", "fa-list-check", short_name="Ass", group_name="Process"),
                "active":    s("Active",    "Ongoing",  "orange", "task", "fa-spinner", short_name="Act", group_name="Process"),
                "completed": s("Completed", "Done",     "green",  "task", "fa-circle-check", short_name="Comp", group_name="Process"),
                "cancelled": s("Cancelled", "Cancelled", "grey",   "task", "fa-ban", short_name="Canc", group_name="Process"),
            },
            "variant": {
                "not_reported":  s("Not reported",                 "Not reported",               "red",    "variant", "fa-circle-question", short_name="NRep", group_name="Process"),
                "reported":      s("Reported",                     "Reported",                   "green",  "variant", "fa-circle-check", short_name="Rep", group_name="Process"),
                "causative":     s("Causative",                    "Causative",                  "green",  "variant", "fa-dna", short_name="Caus", group_name="Causativity", connected_class_keys=("individual",)),
                "suspected":     s("Suspected Causative",          "Suspected causative",        "yellow", "variant", "fa-question", short_name="SCaus", group_name="Causativity"),
                "secondary":     s("Secondary Finding",            "Secondary finding",          "blue",   "variant", "fa-circle-half-stroke", short_name="2nd", group_name="Causativity"),
                "previous":      s("Previously reported",          "Previously reported",        "pink",   "variant", "fa-clock-rotate-left", short_name="PrevRep", group_name="Previous"),
                "ruled_out":     s("Ruled Out",                    "Ruled out",                  "red",    "variant", "fa-xmark", short_name="R/O", group_name="Validity"),
                "ongoing_sanger": s("Ongoing Sanger Confirmation", "Ongoing Sanger confirmation","purple", "variant", "fa-vial", short_name="Sang", group_name="Validity", connected_class_keys=("individual",)),
                "ongoing_func":   s("Ongoing Functional Study",    "Ongoing functional study",   "blue",   "variant", "fa-flask", short_name="Func", group_name="Functional", connected_class_keys=("individual",)),
                "novel_gene":     s("Novel Gene Disease Association","Novel gene-disease association","green","variant","fa-plus", short_name="Novel", group_name="Novel", connected_class_keys=("individual",)),
                "candidate":      s("Candidate Gene-Variant Association","Candidate gene-variant association","yellow","variant","fa-magnifying-glass", short_name="Cand", group_name="Candidate", connected_class_keys=("individual",)),
            },
        }

        id_types = {}
        if not self.dry_run:
            for name in ("RareBoost", "Biobank"):
                example = identifier_type_example_for_name(name)
                id_type, _ = IdentifierType.objects.get_or_create(
                    name=name,
                    defaults={
                        "description": f"{name} identifier",
                        "example": example,
                        "created_by": self.admin_user,
                    },
                )
                if not id_type.example:
                    id_type.example = example
                    id_type.save(update_fields=["example"])
                id_types[name] = id_type

            Institution.objects.get_or_create(
                name="Unknown",
                defaults={"contact": "Placeholder", "created_by": self.admin_user},
            )
            call_command("ozbek_set_id_priorities")

        return statuses, id_types

    def _apply_hpo_terms_to_individual(self, individual, hpo_terms, hpo_source=None) -> None:
        """Attach HPO terms and update affectedness when the import field is explicit."""
        affected = self.statuses["individual"].get("affected")
        healthy = self.statuses["individual"].get("healthy")
        if hpo_terms:
            individual.hpo_terms.add(*hpo_terms)
            if not individual.is_affected:
                individual.is_affected = True
                individual.save(update_fields=["is_affected"])
            if healthy and individual.statuses.filter(pk=healthy.pk).exists():
                individual.statuses.remove(healthy)
            if affected and not individual.statuses.filter(pk=affected.pk).exists():
                individual.statuses.add(affected)
            return

        if str(hpo_source or "").strip().casefold() == "ss":
            if individual.is_affected:
                individual.is_affected = False
                individual.save(update_fields=["is_affected"])
            if affected and individual.statuses.filter(pk=affected.pk).exists():
                individual.statuses.remove(affected)
            if healthy and not individual.statuses.filter(pk=healthy.pk).exists():
                individual.statuses.add(healthy)

    # ------------------------------------------------------------------
    # Workbook helpers
    # ------------------------------------------------------------------

    def _content_types(self) -> dict:
        return {
            "individual": ContentType.objects.get_for_model(Individual),
            "sample":     ContentType.objects.get_for_model(Sample),
            "test":       ContentType.objects.get_for_model(Test),
            "pipeline":   ContentType.objects.get_for_model(Pipeline),
            "analysis":   ContentType.objects.get_for_model(Analysis),
            "analysisreport": ContentType.objects.get_for_model(AnalysisReport),
            "project":    ContentType.objects.get_for_model(Project),
            "task":       ContentType.objects.get_for_model(Task),
            "variant":    ContentType.objects.get_for_model(Variant),
        }

    def _analysis_status_for_import_value(self, value):
        normalized = re.sub(r"\s+", " ", str(value or "").strip().lower())
        if not normalized:
            return None

        status_name = ANALYSIS_IMPORT_STATUS_NAMES.get(normalized)
        if not status_name:
            return None

        cached_status = next(
            (
                status
                for status in self.statuses.get("analysis", {}).values()
                if status.name.lower() == status_name.lower()
            ),
            None,
        )
        if cached_status:
            return cached_status

        return Status.objects.filter(
            content_type=self._content_types()["analysis"],
            name__iexact=status_name,
        ).first()

    def _load_kurumlar_map(self, wb) -> dict:
        """Load institution metadata from the Kurumlar sheet (replaces Gönderen Kurum Harita)."""
        kurumlar_map: dict = {}
        try:
            ws = wb["Kurumlar"]
        except KeyError:
            self.stdout.write(self.style.WARNING(
                "Sheet 'Kurumlar' not found — proceeding without institution metadata."))
            self._record_issue(
                step="step2",
                sheet="Kurumlar",
                severity="warning",
                reason="Sheet not found; institution metadata skipped.",
            )
            return kurumlar_map
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None for c in row):
                continue
            d = dict(zip(headers, row[:len(headers)]))
            name = str(d.get("Kurum") or "").strip()
            if not name:
                continue
            info: dict = {}
            coord_val = d.get("Harita")
            if coord_val:
                try:
                    lat_str, lon_str = [p.strip() for p in str(coord_val).split(",")[:2]]
                    info["coords"] = (float(lat_str), float(lon_str))
                except Exception:
                    pass
            for col, key in (("Şehir", "city"), ("Resmi Ad", "official_name"),
                              ("Merkez", "center_name"), ("Birim", "speciality")):
                val = d.get(col)
                if val:
                    info[key] = str(val).strip()
            kurumlar_map[name] = info
        return kurumlar_map

    def _load_ozbek_lab_rows(self, wb) -> list:
        ws = wb["OZBEK LAB"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None for c in row):
                continue
            rows.append({h: v for h, v in zip(headers, row) if h is not None})
        return rows

    def _ws_rows(self, wb, sheet_name: str):
        """Yield (row_dict, headers) for each non-blank row in a sheet. Returns [] if absent."""
        try:
            ws = wb[sheet_name]
        except KeyError:
            self.stdout.write(self.style.WARNING(
                f"  Sheet '{sheet_name}' not found — skipping."))
            return
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None for c in row):
                continue
            yield dict(zip(headers, row[:len(headers)]))

    # ==================================================================
    # Step 2 — Families + Institutions
    # ==================================================================

    def _step2_families_institutions(self, rows, kurumlar_map) -> tuple:
        self.stdout.write("Step 2: Families and institutions…")
        unique_fids: set = set()
        inst_contacts: dict = {}  # name → set of contact strings

        for row in rows:
            lab_id = row.get("Özbek Lab. ID")
            if not lab_id:
                continue
            fid = get_family_id(lab_id)
            if fid:
                unique_fids.add(fid)

            clinician_assignments = _build_clinician_assignments(
                row.get("Klinisyen"),
                row.get("İletişim Bilgileri - Mail/telefon?"),
                row.get("İletişim Bilgileri - Telefon/mail?"),
            )

            inst_raw = str(row.get("Gönderen Kurum/Birim") or "")
            for name in [n.strip() for n in inst_raw.split(",") if n.strip()]:
                inst_contacts.setdefault(name, set())
                for _, contact_values in clinician_assignments:
                    for contact_value in contact_values:
                        inst_contacts[name].add(contact_value)

        families: dict = {}
        institutions: dict = {}
        unknown_inst = None

        if self.dry_run:
            self.stdout.write(f"  [DRY] {len(unique_fids)} families, "
                              f"{len(inst_contacts)} institutions.")
            return families, institutions, unknown_inst

        for fid in unique_fids:
            family, _ = Family.objects.get_or_create(
                family_id=fid, defaults={"created_by": self.admin_user})
            families[fid] = family

        unknown_inst, _ = Institution.objects.get_or_create(
            name="Unknown",
            defaults={"contact": "Placeholder", "created_by": self.admin_user})

        for name, contacts in inst_contacts.items():
            info = kurumlar_map.get(name, {})
            lat, lon = info.get("coords", (None, None)) if "coords" in info else (None, None)
            institution, created = Institution.objects.get_or_create(
                name=name,
                defaults={
                    "contact": "\n".join(contacts) if contacts else "",
                    "created_by": self.admin_user,
                    "latitude": lat or 0.0,
                    "longitude": lon or 0.0,
                    "city": info.get("city", ""),
                    "official_name": info.get("official_name", ""),
                    "center_name": info.get("center_name", ""),
                    "speciality": info.get("speciality", ""),
                },
            )
            if not created:
                changed = False
                if contacts:
                    existing = set(institution.contact.split("\n")) if institution.contact else set()
                    merged = "\n".join(existing | contacts)
                    if merged != institution.contact:
                        institution.contact = merged; changed = True
                for attr, src in (("latitude", lat), ("longitude", lon),
                                   ("center_name", info.get("center_name")),
                                   ("speciality", info.get("speciality"))):
                    if src and not getattr(institution, attr):
                        setattr(institution, attr, src); changed = True
                if changed:
                    institution.save()
            institutions[name] = institution

        self.stdout.write(f"  Families: {len(families)}  Institutions: {len(institutions)}")
        return families, institutions, unknown_inst

    # ==================================================================
    # Step 3 — Individuals + CrossIdentifiers
    # ==================================================================

    def _step3_individuals(self, rows, families, institutions, unknown_inst) -> set:
        self.stdout.write("Step 3: Individuals and cross identifiers…")
        all_individuals: set = set()
        active = self.statuses["individual"].get("active")
        inactive = self.statuses["individual"].get("inactive")
        unsure_import = self.statuses["individual"].get("unsure_import")

        for row in rows:
            lab_id = row.get("Özbek Lab. ID")
            full_name = row.get("Ad-Soyad")
            if not lab_id or not full_name:
                self._record_issue(
                    step="step3",
                    sheet="OZBEK LAB",
                    severity="warning",
                    reason="Missing required individual fields.",
                    lab_id=lab_id,
                    row=row,
                    context={"missing_lab_id": not bool(lab_id), "missing_full_name": not bool(full_name)},
                )
                continue
            fid = get_family_id(lab_id)
            family = families.get(fid)
            if not family and not self.dry_run:
                self.stdout.write(self.style.WARNING(f"  Family missing for {lab_id} — skip"))
                self._record_issue(
                    step="step3",
                    sheet="OZBEK LAB",
                    severity="warning",
                    reason="Family missing for lab ID.",
                    lab_id=lab_id,
                    row=row,
                    context={"family_id": fid},
                )
                continue

            if self.dry_run:
                self.stdout.write(f"  [DRY] {get_initials(full_name)} ({lab_id})")
                continue

            inst_raw = str(row.get("Gönderen Kurum/Birim") or "")
            inst_list = [institutions.get(n.strip(), unknown_inst)
                         for n in inst_raw.split(",") if n.strip()]
            inst_list = [i for i in inst_list if i] or ([unknown_inst] if unknown_inst else [])

            lab_id_str = str(lab_id).strip()
            is_index = lab_id_str.endswith(".1")

            tc_val = row.get("TC Kimlik No")
            if isinstance(tc_val, float):
                tc_val = int(tc_val)
            elif isinstance(tc_val, str):
                try:
                    tc_val = int(tc_val.strip()) if tc_val.strip() else None
                except ValueError:
                    tc_val = None
            else:
                tc_val = None

            sex_val         = normalize_sex(row.get("Cinsiyet"))
            is_alive_val    = to_bool(row.get("Yaşıyor mu?"))
            age_of_onset    = str(row.get("Age of Onset") or "").strip()
            council_date    = parse_date(row.get("Konsey Tarihi"))
            diagnosis       = str(row.get("Tanı") or "").strip()
            diagnosis_date  = parse_date(row.get("Tanı Tarihi"))
            birth_date      = parse_date(row.get("Doğum Tarihi"))
            icd11_code      = str(row.get("ICD11") or "").strip()
            consanguinity_raw = row.get("Akrabalık")
            consanguinity   = normalize_consanguinity_value(consanguinity_raw)
            registration_date = parse_date(row.get("Geliş Tarihi"))

            individual = find_individual_by_rareboost_id(lab_id)
            if not individual and family:
                individual = Individual.objects.filter(
                    full_name=full_name, family=family).first()

            if not individual:
                individual = Individual.objects.create(
                    full_name=full_name, family=family,
                    birth_date=birth_date, icd11_code=icd11_code,
                    is_index=is_index, tc_identity=tc_val,
                    sex=sex_val or "",
                    is_alive=True if is_alive_val is None else is_alive_val,
                    is_affected=False,
                    age_of_onset=age_of_onset, council_date=council_date,
                    diagnosis=diagnosis, diagnosis_date=diagnosis_date,
                    registration_date=registration_date,
                    created_by=self.admin_user,
                )
                created_statuses = [status for status in (active, unsure_import) if status]
                if created_statuses:
                    individual.statuses.set(created_statuses)
                self.stdout.write(self.style.SUCCESS(
                    f"  Created: {get_initials(full_name)} ({lab_id})"))
            else:
                changed = False
                for attr, val in (
                    ("birth_date", birth_date), ("icd11_code", icd11_code),
                    ("tc_identity", tc_val), ("age_of_onset", age_of_onset),
                    ("council_date", council_date), ("diagnosis", diagnosis),
                    ("diagnosis_date", diagnosis_date),
                    ("registration_date", registration_date),
                ):
                    if val and not getattr(individual, attr):
                        setattr(individual, attr, val); changed = True
                if sex_val and not individual.sex:
                    individual.sex = sex_val; changed = True
                if is_alive_val is not None and individual.is_alive != is_alive_val:
                    individual.is_alive = is_alive_val; changed = True
                if individual.is_index != is_index:
                    individual.is_index = is_index; changed = True
                if changed:
                    individual.save()

            if active and not individual.statuses.filter(pk=active.pk).exists():
                individual.statuses.add(active)
            if inactive and individual.statuses.filter(pk=inactive.pk).exists():
                individual.statuses.remove(inactive)
            if unsure_import and not individual.statuses.filter(pk=unsure_import.pk).exists():
                individual.statuses.add(unsure_import)

            if inst_list:
                individual.institution.set(inst_list)

            if str(consanguinity_raw or "").strip() and family and family.is_consanguineous != consanguinity:
                family.is_consanguineous = consanguinity
                family.save()

            clinician_assignments = _build_clinician_assignments(
                row.get("Klinisyen"),
                row.get("İletişim Bilgileri - Mail/telefon?"),
                row.get("İletişim Bilgileri - Telefon/mail?"),
            )
            for clinician_name, contact_values in clinician_assignments:
                physician = get_or_create_contact(clinician_name, self.admin_user)
                self._link_physician_to_individual_and_institutions(
                    individual,
                    physician,
                    inst_list,
                )
                self._apply_contact_details(physician, contact_values)

            hpo_source = row.get("HPO kodları")
            hpo_terms = get_hpo_terms(hpo_source, self.stdout)
            self._apply_hpo_terms_to_individual(individual, hpo_terms, hpo_source)

            # Projects
            projects_field = row.get("Projeler")
            if projects_field:
                for pname in [p.strip() for p in str(projects_field).split(",") if p.strip()]:
                    project, _ = Project.objects.get_or_create(
                        name=pname,
                        defaults={"created_by": self.admin_user, "priority": "medium"})
                    project.individuals.add(individual)

            # Notes
            parse_and_add_notes(row.get("Kurum Notları"), individual, self.admin_user)
            parse_and_add_notes(row.get("Takip Notları"), individual, self.admin_user)
            parse_and_add_notes(row.get("Genel Notlar/Sonuçlar"), individual, self.admin_user)
            parse_and_add_notes(row.get("İleri tetkik / planlanan"), individual, self.admin_user)
            tamamlanan = row.get("Tamamlanan Tetkik")
            if tamamlanan:
                parse_and_add_notes(
                    f"Tamamlanan tetkikler\n{tamamlanan}", individual, self.admin_user)
                self._import_completed_tests_from_field(
                    individual,
                    tamamlanan,
                    step="step3",
                    sheet="OZBEK LAB",
                    lab_id=lab_id,
                    row=row,
                )

            # CrossIdentifiers
            rb_type = self.id_types.get("RareBoost")
            if rb_type:
                CrossIdentifier.objects.get_or_create(
                    individual=individual, id_type=rb_type,
                    defaults={"id_value": str(lab_id), "created_by": self.admin_user})
            bb_type = self.id_types.get("Biobank")
            bb_val = row.get("Biyobanka ID")
            if bb_type and bb_val:
                CrossIdentifier.objects.get_or_create(
                    individual=individual, id_type=bb_type,
                    defaults={"id_value": str(bb_val), "created_by": self.admin_user})

            # Other IDs column: "IDType:IDValue, IDType2:IDValue2"
            self._parse_other_ids(row.get("Other IDs"), individual)

            all_individuals.add(individual)

        self.stdout.write(f"  Individuals: {len(all_individuals)}")
        return all_individuals

    def _parse_other_ids(self, raw, individual):
        if not raw:
            return
        for pair in str(raw).split(","):
            pair = pair.strip()
            if ":" not in pair:
                continue
            id_type_name, id_value = pair.split(":", 1)
            id_type_name = id_type_name.strip()
            id_value = id_value.strip()
            if not id_type_name or not id_value:
                continue
            example = identifier_type_example_for_name(id_type_name)
            id_type, _ = IdentifierType.objects.get_or_create(
                name=id_type_name,
                defaults={
                    "description": f"{id_type_name} identifier",
                    "example": example,
                    "created_by": self.admin_user,
                })
            if example and not id_type.example:
                id_type.example = example
                id_type.save(update_fields=["example"])
            CrossIdentifier.objects.get_or_create(
                individual=individual, id_type=id_type,
                defaults={"id_value": id_value, "created_by": self.admin_user})

    def _import_completed_tests_from_field(
        self,
        individual,
        raw,
        sample=None,
        step="",
        sheet="",
        lab_id=None,
        row=None,
    ) -> None:
        completed_status = self.statuses["test"].get("completed")
        if not raw or not completed_status:
            return

        target_sample = sample or individual.samples.first() or self._get_placeholder_sample(individual)
        for test_name in self._normalize_test_tokens(str(raw)):
            test_type = get_or_create_test_type(test_name, self.admin_user)
            self._backfill_testtype_report_fields(test_type)
            test, _ = Test.objects.get_or_create(
                sample=target_sample,
                test_type=test_type,
                defaults={"created_by": self.admin_user},
            )
            if not test.statuses.filter(pk=completed_status.pk).exists():
                test.statuses.add(completed_status)

    # ==================================================================
    # Step 4 — Samples
    # ==================================================================

    def _step4_samples(self, rows) -> None:
        self.stdout.write("Step 4: Samples…")
        available = self.statuses["sample"]["available"]
        not_available = self.statuses["sample"]["not_available"]
        planned = self.statuses["sample"]["planned"]
        received = self.statuses["sample"]["received"]
        isolated = self.statuses["sample"]["isolated"]

        for row in rows:
            lab_id = row.get("Özbek Lab. ID")
            if not lab_id:
                self._record_issue(
                    step="step4",
                    sheet="OZBEK LAB",
                    severity="warning",
                    reason="Missing lab ID; samples skipped.",
                    row=row,
                )
                continue
            individual = find_individual_by_rareboost_id(lab_id)
            if not individual:
                self._record_issue(
                    step="step4",
                    sheet="OZBEK LAB",
                    severity="warning",
                    reason="Individual not found; samples skipped.",
                    lab_id=lab_id,
                    row=row,
                )
                continue

            sample_types_raw = str(row.get("Örnek Tipi") or "")
            for sample_type_name in [s.strip() for s in sample_types_raw.split(",") if s.strip()]:
                if self.dry_run:
                    self.stdout.write(f"  [DRY] Sample {sample_type_name} for {lab_id}")
                    continue

                sample_type = get_or_create_sample_type(sample_type_name, self.admin_user)
                # renamed column: "Saklandığı/İzole edildiği yer"
                isolation_by = get_or_create_contact(
                    row.get("Saklandığı/İzole edildiği yer"), self.admin_user)
                receipt_date = parse_date(row.get("Geliş Tarihi"))  # same col as registration_date

                if sample_type_name in ("Tam Kan", "Tam Kan/Serum"):
                    existing = Sample.objects.filter(
                        individual=individual,
                        sample_type__name__in=("Tam Kan", "Tam Kan/Serum")).first()
                else:
                    existing = Sample.objects.filter(
                        individual=individual, sample_type=sample_type).first()

                measurements_raw = row.get("Örnek gön.& OD değ.")
                measurements = str(measurements_raw or "").strip()

                if existing:
                    changed = False
                    if receipt_date and not existing.receipt_date:
                        existing.receipt_date = receipt_date; changed = True
                    if measurements and not existing.sample_measurements:
                        existing.sample_measurements = measurements; changed = True
                    if not existing.isolation_by_id:
                        existing.isolation_by = isolation_by; changed = True
                    if changed:
                        existing.save()
                    if not existing.statuses.exists():
                        status_set = []
                        if existing.receipt_date:
                            if available:
                                status_set.append(available)
                            if isolated:
                                status_set.append(isolated)
                            elif received:
                                status_set.append(received)
                        else:
                            if not_available:
                                status_set.append(not_available)
                            if planned:
                                status_set.append(planned)
                        if status_set:
                            existing.statuses.set(status_set)
                    sample = existing
                else:
                    sample = Sample.objects.create(
                        individual=individual, sample_type=sample_type,
                        receipt_date=receipt_date,
                        sample_measurements=measurements,
                        isolation_by=isolation_by,
                        created_by=self.admin_user,
                    )
                    if receipt_date:
                        sample_statuses = [status for status in (available, isolated or received) if status]
                    else:
                        sample_statuses = [status for status in (not_available, planned) if status]
                    sample.statuses.set(sample_statuses)

                parse_and_add_notes(row.get("Örnek Notları"), sample, self.admin_user)

    # ==================================================================
    # Step 5 — Analiz Takip → Test + Analysis (pipeline linked later)
    # ==================================================================

    def _step5_analiz_takip(self, wb) -> None:
        self.stdout.write("Step 5: Analiz Takip…")
        try:
            ws = wb["Analiz Takip"]
        except KeyError:
            self.stdout.write(self.style.WARNING("  Sheet 'Analiz Takip' not found."))
            return

        raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        headers = [h for h in raw_headers if h is not None]
        leftover_rows = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None for c in row):
                continue
            d = dict(zip(headers, row[:len(headers)]))
            lab_id = d.get("Özbek Lab. ID")
            if not lab_id:
                leftover_rows.append(d)
                self._record_issue(
                    step="step5",
                    sheet="Analiz Takip",
                    severity="warning",
                    reason="Missing lab ID.",
                    row=d,
                )
                continue

            individual = find_individual_by_rareboost_id(lab_id)
            if not individual:
                self.stdout.write(self.style.WARNING(
                    f"  Analiz Takip: no individual for {lab_id}"))
                leftover_rows.append(d)
                self._record_issue(
                    step="step5",
                    sheet="Analiz Takip",
                    severity="warning",
                    reason="Individual not found for analysis row.",
                    lab_id=lab_id,
                    row=d,
                )
                continue

            # VERİ KAYNAĞI is a single value (strip only)
            tt_name = str(d.get("VERİ KAYNAĞI") or "").strip()
            if not tt_name:
                leftover_rows.append(d)
                self._record_issue(
                    step="step5",
                    sheet="Analiz Takip",
                    severity="warning",
                    reason="Missing VERİ KAYNAĞI.",
                    lab_id=lab_id,
                    row=d,
                )
                continue

            if self.dry_run:
                self.stdout.write(f"  [DRY] Test+Analysis for {lab_id} / {tt_name}")
                continue

            sample = individual.samples.first() or self._get_placeholder_sample(individual)

            data_receipt_date = parse_date(d.get("Data Geliş Tarihi"))

            # Test.performed_by is now FK → Institution
            inst_name = str(d.get("Verinin Geldiği Merkez") or "").strip()
            test_inst = None
            if inst_name:
                test_inst, _ = Institution.objects.get_or_create(
                    name=inst_name,
                    defaults={"created_by": self.admin_user})

            test_type = get_or_create_test_type(tt_name, self.admin_user)
            self._backfill_testtype_report_fields(test_type)
            test, test_created = Test.objects.get_or_create(
                sample=sample, test_type=test_type,
                defaults={
                    "data_receipt_date": data_receipt_date,
                    "performed_by": test_inst,
                    "created_by": self.admin_user,
                },
            )
            if test_created:
                test_status = (
                    self.statuses["test"].get("completed")
                    if data_receipt_date
                    else self.statuses["test"].get("waiting")
                    or self.statuses["test"].get("planned")
                )
                if test_status:
                    test.statuses.set([test_status])
            else:
                if data_receipt_date and not test.data_receipt_date:
                    test.data_receipt_date = data_receipt_date
                    test.save()

            # Notes on test
            parse_and_add_notes(d.get("Data Notları"), test, self.admin_user)
            parse_and_add_notes(d.get("Veri İçeriği"), test, self.admin_user)
            parse_and_add_notes(d.get("Veri Notları"), test, self.admin_user)

            # Performed_by for Analysis (comma-separated names in one column)
            performers = self._parse_analysis_performers(d.get("Analizi Yapan"))

            analiz_tarihi  = parse_date(d.get("Reanaliz bitiş tarihi/ayça bitirdiğinde"))
            analiz_turu    = str(d.get("Analiz Türü") or "").strip()
            analiz_durumu  = str(d.get("Analiz Durumu") or "").strip()
            analysis_type  = get_or_create_analysis_type(analiz_turu, self.admin_user) \
                             if analiz_turu else None

            # Analysis — pipeline=None for now; linked by Gennext step
            analysis = Analysis.objects.create(
                pipeline=None,
                type=analysis_type,
                performed_date=analiz_tarihi,
                created_by=self.admin_user,
            )
            if performers:
                analysis.performed_by.set(performers)

            if analiz_durumu:
                a_st = self._analysis_status_for_import_value(analiz_durumu)
                if a_st:
                    analysis.statuses.set([a_st])
                else:
                    self._record_issue(
                        step="step5",
                        sheet="Analiz Takip",
                        severity="warning",
                        reason="Analysis status was not found for imported Analiz Durumu.",
                        lab_id=lab_id,
                        row=d,
                        context={"analiz_durumu": analiz_durumu},
                    )

            # Notes on Analysis (Test Notları goes HERE, not on Test)
            parse_and_add_notes(d.get("Test Notları"), analysis, self.admin_user)
            plan_note = "\n".join(filter(None, [
                str(d.get("PLAN") or "").strip(),
                str(d.get("ANALİZ STATUS") or "").strip(),
            ]))
            if plan_note:
                parse_and_add_notes(plan_note, analysis, self.admin_user)

            self.analysis_map[(str(lab_id), tt_name)] = analysis

        if leftover_rows:
            self.stdout.write(self.style.WARNING(
                f"  {len(leftover_rows)} Analiz Takip rows could not be matched."))
        self.stdout.write(f"  Analyses created: {len(self.analysis_map)}")

    def _parse_analysis_performers(self, field) -> list:
        """Split the performer column into a list of User objects."""
        names: set = set()
        if field:
            for name in re.split(r"[,\n]", str(field)):
                name = name.strip()
                if name:
                    names.add(name)
        users = []
        for name in sorted(names):
            user = get_or_create_user(name, self.admin_user)
            get_or_create_contact(name, self.admin_user, linked_user=user)
            users.append(user)
        return users

    # ==================================================================
    # Step 6 — RarePipe TSV (external file, legacy)
    # ==================================================================

    def _step6_rarepipe(self, tsv_path_str: str) -> None:
        self.stdout.write(f"Step 6: RarePipe TSV {tsv_path_str}…")
        tsv_path = Path(tsv_path_str)
        if not tsv_path.exists():
            self.stdout.write(self.style.ERROR(f"  Not found: {tsv_path}"))
            self._record_issue(
                step="step6",
                sheet="RarePipe TSV",
                severity="error",
                reason="RarePipe TSV file not found.",
                context={"path": str(tsv_path)},
            )
            return

        rows = []
        with tsv_path.open(newline="", encoding="utf-8") as fh:
            for row in csv.reader(fh, delimiter="\t"):
                if any(row):
                    rows.append([c.strip() for c in row])

        versions = {row[4] for row in rows if len(row) >= 5 and row[4]}
        type_map = {v: get_or_create_pipeline_type("RarePipe", self.admin_user, version=v)
                    for v in versions}
        id_map = build_id_map()
        p_completed = self.statuses["pipeline"].get("completed")
        created = skipped = errors = 0

        for i, row in enumerate(rows, start=1):
            if len(row) < 5:
                self._record_issue(
                    step="step6",
                    sheet="RarePipe TSV",
                    severity="error",
                    reason="Row has fewer than 5 columns.",
                    row={"raw_row": row},
                    context={"row_number": i},
                )
                errors += 1; continue
            filename, output_loc, raw_id, input_loc, version = row[:5]
            if not version:
                self._record_issue(
                    step="step6",
                    sheet="RarePipe TSV",
                    severity="error",
                    reason="Missing pipeline version.",
                    row={"raw_row": row},
                    context={"row_number": i},
                )
                errors += 1; continue
            try:
                performed_date = parse_date_from_filename(filename)
            except ValueError as exc:
                self.stdout.write(self.style.ERROR(f"  Row {i}: {exc}"))
                self._record_issue(
                    step="step6",
                    sheet="RarePipe TSV",
                    severity="error",
                    reason="Could not parse performed date from filename.",
                    row={"raw_row": row},
                    context={"row_number": i, "error": str(exc)},
                )
                errors += 1; continue

            individual = id_map.get(normalize_id(raw_id))
            if not individual:
                self._record_issue(
                    step="step6",
                    sheet="RarePipe TSV",
                    severity="warning",
                    reason="Individual not found for normalized ID.",
                    lab_id=raw_id,
                    row={"raw_row": row},
                    context={"row_number": i, "normalized_id": normalize_id(raw_id)},
                )
                skipped += 1; continue

            test = (
                Test.objects.filter(sample__individual=individual,
                                    test_type__name__icontains="WGS").order_by("id").first()
                or Test.objects.filter(sample__individual=individual,
                                       test_type__name__icontains="WES").order_by("id").first()
            )
            if not test:
                self._record_issue(
                    step="step6",
                    sheet="RarePipe TSV",
                    severity="warning",
                    reason="No WGS or WES test found for individual.",
                    lab_id=getattr(individual, "primary_id", raw_id),
                    row={"raw_row": row},
                    context={"row_number": i},
                )
                skipped += 1; continue

            pipeline_type = type_map[version]
            if Pipeline.objects.filter(test=test, type=pipeline_type,
                                        performed_date=performed_date,
                                        output_location=output_loc).exists():
                self._record_issue(
                    step="step6",
                    sheet="RarePipe TSV",
                    severity="info",
                    reason="Duplicate pipeline row skipped.",
                    lab_id=getattr(individual, "primary_id", raw_id),
                    row={"raw_row": row},
                    context={"row_number": i},
                )
                skipped += 1; continue

            if self.dry_run:
                created += 1; continue

            pipeline = Pipeline.objects.create(
                test=test, performed_date=performed_date, performed_by=self.admin_user,
                type=pipeline_type, input_location=input_loc, output_location=output_loc,
                created_by=self.admin_user)
            if p_completed:
                pipeline.statuses.set([p_completed])
            Analysis.objects.get_or_create(
                pipeline=pipeline,
                defaults={"created_by": self.admin_user})
            created += 1

        self.stdout.write(self.style.SUCCESS(
            f"  RarePipe TSV: created={created} skipped={skipped} errors={errors}"))

    # ==================================================================
    # Step 7 — Parent links
    # ==================================================================

    def _step7_parent_links(self, families: dict) -> None:
        self.stdout.write("Step 7: Parent links…")
        updated = 0
        for fam_obj in families.values():
            code_to_ind: dict = {}
            for ind in fam_obj.individuals.all():
                lab_id_val = ind.primary_id
                if not lab_id_val or str(lab_id_val).upper().startswith("NO "):
                    continue
                parts = str(lab_id_val).split(".")
                if len(parts) < 2:
                    continue
                code = ".".join(parts[1:])
                code_to_ind.setdefault(code, ind)

            mother = code_to_ind.get("2")
            father = code_to_ind.get("3")
            if not mother and not father:
                continue
            for code, child in code_to_ind.items():
                if not (code == "1" or code.startswith("1.")):
                    continue
                changed = False
                if mother and child.mother_id != mother.id and child.id != mother.id:
                    child.mother = mother; changed = True
                if father and child.father_id != father.id and child.id != father.id:
                    child.father = father; changed = True
                if changed:
                    child.save(); updated += 1
        self.stdout.write(self.style.SUCCESS(f"  Updated: {updated}"))

    # ==================================================================
    # Step 8 — Sanger Konfirmasyonları
    # ==================================================================

    def _step_sanger(self, wb) -> None:
        self.stdout.write("Step 8: Sanger Konfirmasyonları…")
        sanger_type = get_or_create_test_type("Sanger", self.admin_user)
        self._backfill_testtype_report_fields(sanger_type)
        created = skipped = 0
        for d in self._ws_rows(wb, "Sanger Konfirmasyonları"):
            lab_id = str(d.get("Özbek Lab. ID") or "").strip()
            if not lab_id:
                self._record_issue(
                    step="step8",
                    sheet="Sanger Konfirmasyonları",
                    severity="warning",
                    reason="Missing lab ID.",
                    row=d,
                )
                skipped += 1; continue
            individual = find_individual_by_rareboost_id(lab_id)
            if not individual:
                self._record_issue(
                    step="step8",
                    sheet="Sanger Konfirmasyonları",
                    severity="warning",
                    reason="Individual not found.",
                    lab_id=lab_id,
                    row=d,
                )
                skipped += 1; continue
            if self.dry_run:
                created += 1; continue

            sample = individual.samples.first() or self._get_placeholder_sample(individual)
            test, test_created = Test.objects.get_or_create(
                sample=sample, test_type=sanger_type,
                defaults={"created_by": self.admin_user})
            if test_created:
                test_completed = self.statuses["test"].get("completed")
                if test_completed:
                    test.statuses.set([test_completed])

            note_lines = "\n".join(filter(None, [
                str(d.get("Chromosomal Position") or "").strip(),
                str(d.get("Sanger Conf. Status") or "").strip(),
            ]))
            parse_and_add_notes(note_lines, test, self.admin_user)
            created += 1
        self.stdout.write(f"  Sanger: created={created} skipped={skipped}")

    # ==================================================================
    # Step 9 — WGS_TÜSEB
    # ==================================================================

    def _step_wgs_tuseb(self, wb) -> None:
        self.stdout.write("Step 9: WGS_TÜSEB…")
        project, _ = Project.objects.get_or_create(
            name="WGS - TÜSEB",
            defaults={"created_by": self.admin_user, "priority": "medium"})
        wgs_type = get_or_create_test_type("WGS", self.admin_user)
        self._backfill_testtype_report_fields(wgs_type)
        completed  = self.statuses["test"].get("completed")
        waiting    = self.statuses["test"].get("waiting")
        id_map = build_id_map()
        bb_type = self.id_types.get("Biobank")

        measurement_cols = [
            "Total Hacim ( ul)", "Nanodrop Ölçümü (ng/ul)",
            "A260/280", "A260/230", "Qubit Ölçümü",
        ]
        created = skipped = 0

        for d in self._ws_rows(wb, "WGS_TÜSEB"):
            non_empty_keys = {
                str(key).strip()
                for key, value in d.items()
                if value is not None and str(value).strip()
            }
            if non_empty_keys and all(key.startswith("Örnek No") for key in non_empty_keys):
                self._record_issue(
                    step="step9",
                    sheet="WGS_TÜSEB",
                    severity="info",
                    reason="Separator row in WGS_TÜSEB skipped.",
                    row=d,
                    context={"non_empty_columns": sorted(non_empty_keys)},
                )
                skipped += 1
                continue

            lab_id = str(d.get("Özbek Lab. ID") or "").strip()
            individual = None
            if lab_id:
                individual = Individual.objects.filter(
                    cross_ids__id_value=lab_id).first()
            if not individual and bb_type:
                bb_val = str(d.get("Biyobanka ID") or "").strip()
                if bb_val:
                    individual = Individual.objects.filter(
                        cross_ids__id_type=bb_type,
                        cross_ids__id_value=bb_val).first()
            if not individual:
                self._record_issue(
                    step="step9",
                    sheet="WGS_TÜSEB",
                    severity="warning",
                    reason="Individual not found by RareBoost or Biobank ID.",
                    lab_id=lab_id or d.get("Biyobanka ID"),
                    row=d,
                )
                skipped += 1; continue
            if self.dry_run:
                created += 1; continue

            project.individuals.add(individual)
            sample = individual.samples.first() or self._get_placeholder_sample(individual)

            # Consolidated measurements
            parts = [f"{col}: {d[col]}"
                     for col in measurement_cols
                     if d.get(col) is not None and str(d.get(col)).strip()]
            measurements = " | ".join(parts)

            test, test_created = Test.objects.get_or_create(
                sample=sample, test_type=wgs_type,
                defaults={
                    "service_send_date": parse_date(d.get("Dizilemeye Gönderilme Tarihi")),
                    "data_receipt_date": parse_date(d.get("Data Gelme Tarihi")),
                    "created_by": self.admin_user,
                })
            if test_created and measurements:
                sample.sample_measurements = (
                    (sample.sample_measurements + " | " if sample.sample_measurements else "")
                    + measurements)
                sample.save()

            # Status
            status_raw = str(d.get("Data Geliş Durumu") or "").strip().lower()
            if status_raw == "geldi":
                test_status = completed
            elif status_raw == "data bekleniyor":
                test_status = waiting
            else:
                test_status = None
            if test_status and not test.statuses.exists():
                test.statuses.set([test_status])

            parse_and_add_notes(d.get("Data Notları"), test, self.admin_user)
            created += 1

        self.stdout.write(f"  WGS_TÜSEB: created={created} skipped={skipped}")

    # ==================================================================
    # Step 10 — External
    # ==================================================================

    def _step_external(self, wb, kurumlar_map) -> None:
        self.stdout.write("Step 10: External individuals…")
        created_count = skipped = 0

        for d in self._ws_rows(wb, "External"):
            id_type_name = str(d.get("ID Type") or "").strip()
            id_value     = str(d.get("ID Value") or "").strip()
            full_name    = str(d.get("Ad-Soyad") or "").strip()
            if not id_type_name or not id_value or not full_name:
                self._record_issue(
                    step="step10",
                    sheet="External",
                    severity="warning",
                    reason="Missing required external individual fields.",
                    lab_id=id_value,
                    row=d,
                    context={
                        "missing_id_type": not bool(id_type_name),
                        "missing_id_value": not bool(id_value),
                        "missing_full_name": not bool(full_name),
                    },
                )
                skipped += 1; continue
            if self.dry_run:
                self.stdout.write(f"  [DRY] External: {full_name}")
                created_count += 1; continue

            # One family per individual (family_id = sanitized id_value)
            fam_id = re.sub(r"[^A-Za-z0-9_\-]", "_", id_value)
            family, _ = Family.objects.get_or_create(
                family_id=fam_id, defaults={"created_by": self.admin_user})

            # Institution
            inst_raw = str(d.get("Gönderen Kurum/Birim") or "")
            inst_list = []
            for name in [n.strip() for n in inst_raw.split(",") if n.strip()]:
                info = kurumlar_map.get(name, {})
                lat, lon = info.get("coords", (None, None)) if "coords" in info else (None, None)
                inst, _ = Institution.objects.get_or_create(
                    name=name,
                    defaults={
                        "contact": "", "created_by": self.admin_user,
                        "latitude": lat or 0.0, "longitude": lon or 0.0,
                        "city": info.get("city", ""), "official_name": info.get("official_name", ""),
                    })
                inst_list.append(inst)

            tc_val = d.get("TC Kimlik No")
            if isinstance(tc_val, float):
                tc_val = int(tc_val)
            elif isinstance(tc_val, str):
                try:
                    tc_val = int(tc_val.strip()) if tc_val.strip() else None
                except ValueError:
                    tc_val = None
            else:
                tc_val = None

            # Cross-ID type
            example = identifier_type_example_for_name(id_type_name)
            ext_id_type, _ = IdentifierType.objects.get_or_create(
                name=id_type_name,
                defaults={
                    "description": id_type_name,
                    "example": example,
                    "created_by": self.admin_user,
                })
            if example and not ext_id_type.example:
                ext_id_type.example = example
                ext_id_type.save(update_fields=["example"])

            individual = Individual.objects.filter(
                cross_ids__id_type=ext_id_type, cross_ids__id_value=id_value).first()

            if not individual:
                individual = Individual.objects.create(
                    full_name=full_name, family=family,
                    birth_date=parse_date(d.get("Doğum Tarihi")),
                    icd11_code=str(d.get("ICD11") or ""),
                    is_index=True, tc_identity=tc_val,
                    council_date=parse_date(d.get("Konsey Tarihi")),
                    created_by=self.admin_user)
                ext_active = self.statuses["individual"].get("active")
                ext_unsure = self.statuses["individual"].get("unsure_import")
                ext_statuses = [status for status in (ext_active, ext_unsure) if status]
                if ext_statuses:
                    individual.statuses.set(ext_statuses)

            if inst_list:
                individual.institution.set(inst_list)

            CrossIdentifier.objects.get_or_create(
                individual=individual, id_type=ext_id_type,
                defaults={"id_value": id_value, "created_by": self.admin_user})

            bb_type = self.id_types.get("Biobank")
            bb_val = str(d.get("Biyobanka ID") or "").strip()
            if bb_type and bb_val:
                CrossIdentifier.objects.get_or_create(
                    individual=individual, id_type=bb_type,
                    defaults={"id_value": bb_val, "created_by": self.admin_user})

            self._parse_other_ids(d.get("Other IDs"), individual)

            clinician_assignments = _build_clinician_assignments(
                d.get("Klinisyen"),
                d.get("İletişim Bilgileri - Mail/telefon?"),
                d.get("İletişim Bilgileri - Telefon/mail?"),
            )
            for clinician_name, contact_values in clinician_assignments:
                physician = get_or_create_contact(clinician_name, self.admin_user)
                self._link_physician_to_individual_and_institutions(
                    individual,
                    physician,
                    inst_list,
                )
                self._apply_contact_details(physician, contact_values)

            hpo_source = d.get("HPO kodları")
            hpo_terms = get_hpo_terms(hpo_source, self.stdout)
            self._apply_hpo_terms_to_individual(individual, hpo_terms, hpo_source)

            # Sample
            receipt_date = parse_date(d.get("Geliş Tarihi/ay/gün/yıl"))
            sample_type_name = str(d.get("Örnek Tipi") or "").strip()
            sample = None
            if sample_type_name:
                sample_type = get_or_create_sample_type(sample_type_name, self.admin_user)
                isolation_by = get_or_create_contact(
                    d.get("İzolasyonu yapan"), self.admin_user)
                measurements = str(d.get("Örnek gön.& OD değ.") or "").strip()
                sample, s_created = Sample.objects.get_or_create(
                    individual=individual, sample_type=sample_type,
                    defaults={
                        "receipt_date": receipt_date,
                        "sample_measurements": measurements,
                        "isolation_by": isolation_by,
                        "created_by": self.admin_user,
                    })
                if s_created:
                    avail = self.statuses["sample"]["available"]
                    not_available = self.statuses["sample"].get("not_available")
                    planned = self.statuses["sample"].get("planned")
                    received = self.statuses["sample"].get("received")
                    isolated = self.statuses["sample"].get("isolated")
                    if receipt_date:
                        sample_statuses = [status for status in (avail, isolated or received) if status]
                    else:
                        sample_statuses = [status for status in (not_available, planned) if status]
                    sample.statuses.set(sample_statuses)
                parse_and_add_notes(d.get("Örnek Notları"), sample, self.admin_user)

            # Test
            test_name = str(d.get("Çalışılan Test Adı") or "").strip()
            if test_name and sample:
                tt = get_or_create_test_type(test_name, self.admin_user)
                self._backfill_testtype_report_fields(tt)
                test, t_created = Test.objects.get_or_create(
                    sample=sample, test_type=tt,
                    defaults={
                        "performed_date": parse_date(d.get("Çalışılma Tarihi")),
                        "service_send_date": parse_date(d.get("Hiz.Alım.Gön. Tarihi")),
                        "data_receipt_date": parse_date(d.get("Data Geliş tarihi")),
                        "created_by": self.admin_user,
                    })
                parse_and_add_notes(d.get("Test Notları"), test, self.admin_user)

            # Notes on individual
            for col in ("Kurum Notları", "Takip Notları",
                        "Genel Notlar/Sonuçlar", "İleri tetkik / planlanan"):
                parse_and_add_notes(d.get(col), individual, self.admin_user)
            tamamlanan = d.get("Tamamlanan Tetkik")
            if tamamlanan:
                parse_and_add_notes(
                    f"Tamamlanan tetkikler\n{tamamlanan}", individual, self.admin_user)
                self._import_completed_tests_from_field(
                    individual,
                    tamamlanan,
                    sample=sample,
                    step="step10",
                    sheet="External",
                    lab_id=lab_id,
                    row=d,
                )

            # Projects
            projects_field = d.get("Projeler")
            if projects_field:
                for pname in [p.strip() for p in str(projects_field).split(",") if p.strip()]:
                    project, _ = Project.objects.get_or_create(
                        name=pname,
                        defaults={"created_by": self.admin_user, "priority": "medium"})
                    project.individuals.add(individual)

            created_count += 1

        self.stdout.write(f"  External: created/updated={created_count} skipped={skipped}")

    # ==================================================================
    # Step 11 — Long Read (Katar / Dubai)
    # ==================================================================

    def _step_long_read(self, wb, sheet_name: str, note_tag: str, project_name: str) -> None:
        self.stdout.write(f"Step 11 ({note_tag}): {sheet_name}…")
        project, _ = Project.objects.get_or_create(
            name=project_name,
            defaults={"created_by": self.admin_user, "priority": "medium"})
        lr_type = get_or_create_test_type("Long Read WGS", self.admin_user)
        self._backfill_testtype_report_fields(lr_type)
        sent_status = self.statuses["test"].get("waiting")
        note_cols = [
            "Hastalık Grubu",
            "Biyobankada mevcut örnek tipleri",
            "Materyal",
            "Kalite Metrikleri",
        ]
        created = skipped = 0
        for d in self._ws_rows(wb, sheet_name):
            lab_id = str(d.get("RareBoost ID") or "").strip()
            if not lab_id:
                self._record_issue(
                    step="step11",
                    sheet=sheet_name,
                    severity="warning",
                    reason="Missing RareBoost ID.",
                    row=d,
                    context={"note_tag": note_tag},
                )
                skipped += 1; continue
            individual = find_individual_by_rareboost_id(lab_id)
            if not individual:
                self._record_issue(
                    step="step11",
                    sheet=sheet_name,
                    severity="warning",
                    reason="Individual not found.",
                    lab_id=lab_id,
                    row=d,
                    context={"note_tag": note_tag},
                )
                skipped += 1; continue
            if self.dry_run:
                created += 1; continue

            project.individuals.add(individual)
            sample = individual.samples.first() or self._get_placeholder_sample(individual)
            test, t_created = Test.objects.get_or_create(
                sample=sample, test_type=lr_type,
                defaults={"created_by": self.admin_user})
            if t_created and sent_status:
                test.statuses.set([sent_status])

            lines = [note_tag]
            for col in note_cols:
                val = str(d.get(col) or "").strip()
                if val:
                    lines.append(f"{col}: {val}")
            parse_and_add_notes("\n".join(lines), test, self.admin_user)
            created += 1
        self.stdout.write(f"  {note_tag}: created={created} skipped={skipped}")

    # ==================================================================
    # Step 12 — CP_COHORT
    # ==================================================================

    def _step_cp_cohort(self, wb) -> None:
        self.stdout.write("Step 12: CP_COHORT…")
        project, _ = Project.objects.get_or_create(
            name="CP Cohort",
            defaults={"created_by": self.admin_user, "priority": "medium"})
        added = 0
        for d in self._ws_rows(wb, "CP_COHORT"):
            lab_id = str(d.get("Özbek Lab. ID") or "").strip()
            if not lab_id:
                self._record_issue(
                    step="step12",
                    sheet="CP_COHORT",
                    severity="warning",
                    reason="Missing lab ID.",
                    row=d,
                )
                continue
            individual = find_individual_by_rareboost_id(lab_id)
            if not individual:
                self._record_issue(
                    step="step12",
                    sheet="CP_COHORT",
                    severity="warning",
                    reason="Individual not found.",
                    lab_id=lab_id,
                    row=d,
                )
                continue
            if not self.dry_run:
                project.individuals.add(individual)
                parse_and_add_notes(d.get("Analiz Sonucu"), individual, self.admin_user)
            added += 1
        self.stdout.write(f"  CP_COHORT: {added} individuals")

    # ==================================================================
    # Step 13 — RNA SEQ
    # ==================================================================

    def _step_rna_seq(self, wb) -> None:
        self.stdout.write("Step 13: RNA SEQ…")
        rna_type = get_or_create_test_type("RNA Seq", self.admin_user)
        self._backfill_testtype_report_fields(rna_type)
        created = skipped = 0
        for d in self._ws_rows(wb, "RNA SEQ"):
            lab_id = str(d.get("Özbek Lab. ID") or "").strip()
            if not lab_id:
                self._record_issue(
                    step="step13",
                    sheet="RNA SEQ",
                    severity="warning",
                    reason="Missing lab ID.",
                    row=d,
                )
                skipped += 1; continue
            individual = find_individual_by_rareboost_id(lab_id)
            if not individual:
                self._record_issue(
                    step="step13",
                    sheet="RNA SEQ",
                    severity="warning",
                    reason="Individual not found.",
                    lab_id=lab_id,
                    row=d,
                )
                skipped += 1; continue
            if self.dry_run:
                created += 1; continue

            sample = individual.samples.first() or self._get_placeholder_sample(individual)
            test, _ = Test.objects.get_or_create(
                sample=sample, test_type=rna_type,
                defaults={
                    "service_send_date": parse_date(d.get("Dizilemeye Gönderim Tarihi")),
                    "created_by": self.admin_user,
                })
            status_note = str(d.get(
                "Data Yüklenme Tarihi (G&More)/Status") or "").strip()
            if status_note:
                parse_and_add_notes(
                    f"Data Yüklenme Tarihi (G&More)/Status: {status_note}",
                    test, self.admin_user)
            parse_and_add_notes(d.get("Notlar"), test, self.admin_user)
            created += 1
        self.stdout.write(f"  RNA SEQ: created={created} skipped={skipped}")

    # ==================================================================
    # Step 14 — Gennext Analiz Listesi (creates Pipelines, links analyses)
    # ==================================================================

    def _step_gennext_analiz(self, wb) -> None:
        self.stdout.write("Step 14: Gennext Analiz Listesi…")
        gennext_type = get_or_create_pipeline_type("Gennext", self.admin_user)
        p_completed  = self.statuses["pipeline"].get("completed")
        id_map = build_id_map()
        created = skipped = errors = 0

        for d in self._ws_rows(wb, "Gennext Analiz Listesi"):
            lab_id = str(d.get("Gennext ID") or "").strip()
            if not lab_id:
                self._record_issue(
                    step="step14",
                    sheet="Gennext Analiz Listesi",
                    severity="warning",
                    reason="Missing Gennext ID.",
                    row=d,
                )
                skipped += 1; continue
            performed_date = parse_date(d.get("Gennext Date"))
            if not performed_date:
                self.stdout.write(self.style.WARNING(
                    f"  Gennext: no date for {lab_id} — skipping"))
                self._record_issue(
                    step="step14",
                    sheet="Gennext Analiz Listesi",
                    severity="warning",
                    reason="Missing or invalid Gennext Date.",
                    lab_id=lab_id,
                    row=d,
                )
                skipped += 1; continue

            individual = id_map.get(normalize_id(lab_id))
            if not individual:
                self._record_issue(
                    step="step14",
                    sheet="Gennext Analiz Listesi",
                    severity="warning",
                    reason="Individual not found by any cross identifier.",
                    lab_id=lab_id,
                    row=d,
                )
                skipped += 1; continue

            test = (
                Test.objects.filter(sample__individual=individual,
                                    test_type__name__icontains="WES").order_by("id").first()
                or Test.objects.filter(sample__individual=individual,
                                       test_type__name__icontains="WGS").order_by("id").first()
            )
            if not test:
                self.stdout.write(self.style.WARNING(
                    f"  Gennext: no WES/WGS test for {lab_id} — creating fallback WES test"))
                wes_tt = get_or_create_test_type("WES", self.admin_user)
                self._backfill_testtype_report_fields(wes_tt)
                sample = individual.samples.first() or self._get_placeholder_sample(individual)
                test = Test.objects.create(
                    sample=sample,
                    test_type=wes_tt,
                    created_by=self.admin_user,
                )
                synthetic_statuses = [
                    self.statuses["test"].get("previous"),
                    self.statuses["test"].get("unsure_import"),
                ]
                test.statuses.set([s for s in synthetic_statuses if s])
                self._record_issue(
                    step="step14",
                    sheet="Gennext Analiz Listesi",
                    severity="info",
                    reason="No WES or WGS test found; created fallback WES test with Previous + Unsure Import.",
                    lab_id=lab_id,
                    row=d,
                )

            output_loc = ""
            if Pipeline.objects.filter(test=test, type=gennext_type,
                                        performed_date=performed_date).exists():
                self._record_issue(
                    step="step14",
                    sheet="Gennext Analiz Listesi",
                    severity="info",
                    reason="Duplicate pipeline already exists.",
                    lab_id=lab_id,
                    row=d,
                )
                skipped += 1; continue

            if self.dry_run:
                created += 1; continue

            try:
                pipeline = Pipeline.objects.create(
                    test=test, performed_date=performed_date,
                    performed_by=self.admin_user,
                    type=gennext_type, output_location=output_loc,
                    created_by=self.admin_user)
                if p_completed:
                    pipeline.statuses.set([p_completed])

                note_lines = []
                gennext_note = str(d.get("Gennext") or "").strip()
                if gennext_note:
                    note_lines.append(f"Gennext: {gennext_note}")
                gennext_hash = str(d.get("Gennext Hash") or "").strip()
                if gennext_hash:
                    note_lines.append(f"Gennext Hash: {gennext_hash}")
                if note_lines:
                    parse_and_add_notes("\n".join(note_lines), pipeline, self.admin_user)

                # Link an unlinked Analysis from analysis_map for this individual
                tt_name = test.test_type.name
                analysis = self.analysis_map.get((str(lab_id), tt_name))
                if analysis and analysis.pipeline_id is None:
                    analysis.pipeline = pipeline
                    analysis.save()

                created += 1
            except Exception as exc:
                self.stdout.write(self.style.ERROR(f"  Gennext error {lab_id}: {exc}"))
                self._record_issue(
                    step="step14",
                    sheet="Gennext Analiz Listesi",
                    severity="error",
                    reason="Unhandled error while creating Gennext pipeline.",
                    lab_id=lab_id,
                    row=d,
                    context={"error": str(exc)},
                )
                errors += 1

        self.stdout.write(self.style.SUCCESS(
            f"  Gennext: created={created} skipped={skipped} errors={errors}"))

    # ==================================================================
    # Step 15 — RarePipe Analiz Listesi
    # ==================================================================

    def _step_rarepipe_analiz(self, wb) -> None:
        self.stdout.write("Step 15: RarePipe Analiz Listesi…")
        rarepipe_type = get_or_create_pipeline_type("RarePipe", self.admin_user)
        p_completed = self.statuses["pipeline"].get("completed")
        id_map = build_id_map()
        created = skipped = errors = 0

        for d in self._ws_rows(wb, "RarePipe Analiz Listesi"):
            sheet_name = str(d.get("Samplesheet Name") or "").strip()
            performed_date = parse_date(d.get("Date"))
            sample_id = str(
                d.get("Sample ID (Note)")
                or d.get("Sample ID")
                or ""
            ).strip()
            matched_id = str(
                d.get("Matched ID")
                or d.get("Matching Sample ID")
                or d.get("ID")
                or ""
            ).strip()

            if not performed_date:
                self._record_issue(
                    step="step15",
                    sheet="RarePipe Analiz Listesi",
                    severity="warning",
                    reason="Missing or invalid Date.",
                    row=d,
                )
                skipped += 1
                continue

            if not matched_id:
                self._record_issue(
                    step="step15",
                    sheet="RarePipe Analiz Listesi",
                    severity="warning",
                    reason="Missing Matched ID.",
                    row=d,
                )
                skipped += 1
                continue

            individual = id_map.get(normalize_id(matched_id))
            if not individual:
                self._record_issue(
                    step="step15",
                    sheet="RarePipe Analiz Listesi",
                    severity="warning",
                    reason="Individual not found by any cross identifier.",
                    lab_id=matched_id,
                    row=d,
                )
                skipped += 1
                continue

            test = (
                Test.objects.filter(sample__individual=individual,
                                    test_type__name__icontains="WGS").order_by("id").first()
                or Test.objects.filter(sample__individual=individual,
                                       test_type__name__icontains="WES").order_by("id").first()
            )
            if not test:
                if self.dry_run:
                    self._record_issue(
                        step="step15",
                        sheet="RarePipe Analiz Listesi",
                        severity="info",
                        reason="No WGS or WES test found; dry-run would create fallback WES test with Previous + Unsure Import.",
                        lab_id=matched_id,
                        row=d,
                    )
                    skipped += 1
                    continue

                wes_tt = get_or_create_test_type("WES", self.admin_user)
                self._backfill_testtype_report_fields(wes_tt)
                sample = individual.samples.first() or self._get_placeholder_sample(individual)
                test = Test.objects.create(
                    sample=sample,
                    test_type=wes_tt,
                    created_by=self.admin_user,
                )
                synthetic_statuses = [
                    self.statuses["test"].get("previous"),
                    self.statuses["test"].get("unsure_import"),
                ]
                test.statuses.set([s for s in synthetic_statuses if s])
                self._record_issue(
                    step="step15",
                    sheet="RarePipe Analiz Listesi",
                    severity="info",
                    reason="No WGS or WES test found; created fallback WES test with Previous + Unsure Import.",
                    lab_id=matched_id,
                    row=d,
                    context={"test_type": "WES"},
                )

            note_lines = [line for line in (
                f"Samplesheet Name: {sheet_name}" if sheet_name else "",
                f"Sample ID: {sample_id}" if sample_id else "",
            ) if line]

            if Pipeline.objects.filter(
                test=test,
                type=rarepipe_type,
                performed_date=performed_date,
            ).exists():
                self._record_issue(
                    step="step15",
                    sheet="RarePipe Analiz Listesi",
                    severity="info",
                    reason="Duplicate RarePipe pipeline already exists.",
                    lab_id=matched_id,
                    row=d,
                )
                skipped += 1
                continue

            if self.dry_run:
                created += 1
                continue

            try:
                pipeline = Pipeline.objects.create(
                    test=test,
                    performed_date=performed_date,
                    performed_by=self.admin_user,
                    type=rarepipe_type,
                    created_by=self.admin_user,
                )
                if p_completed:
                    pipeline.statuses.set([p_completed])

                if note_lines:
                    parse_and_add_notes("\n".join(note_lines), pipeline, self.admin_user)

                Analysis.objects.get_or_create(
                    pipeline=pipeline,
                    defaults={"created_by": self.admin_user},
                )
                created += 1
            except Exception as exc:
                self._record_issue(
                    step="step15",
                    sheet="RarePipe Analiz Listesi",
                    severity="error",
                    reason="Unhandled error while importing RarePipe pipeline.",
                    lab_id=matched_id,
                    row=d,
                    context={"error": str(exc)},
                )
                errors += 1

        self.stdout.write(self.style.SUCCESS(
            f"  RarePipe Analiz Listesi: created={created} skipped={skipped} errors={errors}"))

    # ==================================================================
    # Step 16 — Variant List
    # ==================================================================

    def _step_variants(self, wb) -> None:
        self.stdout.write("Step 16: Variant List…")
        variant_ws = None
        for sheet_name in wb.sheetnames:
            if sheet_name == "Variant List":
                variant_ws = wb[sheet_name]
                break
        if not variant_ws:
            self.stdout.write(self.style.WARNING("  Sheet 'Variant List' not found."))
            self._record_issue(
                step="step16",
                sheet="Variant List",
                severity="warning",
                reason="Sheet not found.",
            )
            return

        variant_ct = ContentType.objects.get_for_model(Variant)
        headers = [c.value for c in next(variant_ws.iter_rows(min_row=1, max_row=1))]
        imported = skipped = errors = 0

        for row in variant_ws.iter_rows(min_row=2, values_only=True):
            if all(c is None for c in row):
                continue
            d = dict(zip(headers, row))
            lab_id = str(d.get("Özbek Lab. ID") or "").strip()
            if not lab_id:
                self._record_issue(
                    step="step16",
                    sheet="Variant List",
                    severity="warning",
                    reason="Missing lab ID.",
                    row=d,
                )
                continue

            individual = find_individual_by_rareboost_id(lab_id)
            if not individual:
                self._record_issue(
                    step="step16",
                    sheet="Variant List",
                    severity="warning",
                    reason="Individual not found.",
                    lab_id=lab_id,
                    row=d,
                )
                skipped += 1; continue

            # Zygosity — strict mapping, skip if unrecognised
            zyg = _map_zygosity_strict(
                d.get("Zygosity"),
                warn_fn=lambda msg: self.stdout.write(self.style.WARNING(msg)))
            if zyg is None:
                self._record_issue(
                    step="step16",
                    sheet="Variant List",
                    severity="warning",
                    reason="Invalid or empty zygosity; variant skipped.",
                    lab_id=lab_id,
                    row=d,
                )
                errors += 1; continue

            parsed = parse_variant_string(d.get("Chromosomal Position"))
            if not parsed:
                # ⚠ CNV/SV/Repeat formats not yet implemented (Q5 — ask after implementation)
                self.stdout.write(self.style.WARNING(
                    f"  Cannot parse '{d.get('Chromosomal Position')}' for {lab_id} "
                    f"— CNV/SV/Repeat formats not yet implemented."))
                self._record_issue(
                    step="step16",
                    sheet="Variant List",
                    severity="warning",
                    reason="Chromosomal Position could not be parsed; CNV/SV/Repeat formats not implemented.",
                    lab_id=lab_id,
                    row=d,
                )
                errors += 1; continue

            chrom, start_str, ref, alt = parsed

            if self.dry_run:
                imported += 1; continue

            # First analysis for this individual (no Veri Kaynağı mapping available)
            analysis = Analysis.objects.filter(
                pipeline__test__sample__individual=individual).order_by("id").first()
            # Fallback: check analysis_map
            if analysis is None:
                for (lid, _), a in self.analysis_map.items():
                    if lid == str(lab_id):
                        analysis = a; break

            try:
                snv, created_snv = SNV.objects.get_or_create(
                    individual=individual,
                    chromosome=chrom,
                    start=int(start_str),
                    reference=ref,
                    alternate=alt,
                    defaults={
                        "end": int(start_str),
                        "zygosity": zyg,
                        "analysis": analysis,
                        "created_by": self.admin_user,
                    })
                if not created_snv and analysis and snv.analysis_id != analysis.id:
                    snv.analysis = analysis; snv.save()

                # Statuses are seeded in step 1 and matched by name here.
                varyant_durumu = str(d.get("Statuses") or "").strip()
                if varyant_durumu and created_snv:
                    v_st = Status.objects.filter(
                        content_type=variant_ct,
                        name__iexact=varyant_durumu,
                    ).first() or get_or_create_status(
                        varyant_durumu, "", "gray", self.admin_user, variant_ct
                    )
                    if v_st:
                        snv.statuses.set([v_st])

                imported += 1
            except Exception as exc:
                self.stdout.write(self.style.ERROR(f"  Variant error {lab_id}: {exc}"))
                self._record_issue(
                    step="step16",
                    sheet="Variant List",
                    severity="error",
                    reason="Unhandled error while importing variant.",
                    lab_id=lab_id,
                    row=d,
                    context={"error": str(exc)},
                )
                errors += 1

        self.stdout.write(self.style.SUCCESS(
            f"  Variants: imported={imported} skipped={skipped} errors={errors}"))

    # ==================================================================
    # Step 18 — File attachments
    # ==================================================================

    def _step_file_attachments(self, forms_dir_str, reports_dir_str) -> None:
        self.stdout.write("Step 18: File attachments…")
        id_regex = re.compile(r"^(?P<lab_id>(?:RB_\d{4}_[\d\.]+|RD3\.F\d+(?:\.\d+)+))")

        if forms_dir_str:
            forms_dir = Path(forms_dir_str)
            if forms_dir.exists():
                for fp in sorted(forms_dir.glob("*")):
                    if not fp.is_file() or fp.name.startswith("."): continue
                    m = id_regex.match(fp.name)
                    if not m:
                        self._record_issue(
                            step="step18",
                            sheet="forms_dir",
                            severity="warning",
                            reason="Form filename did not match RareBoost ID pattern.",
                            context={"file": fp.name},
                        )
                        continue
                    ind = find_individual_by_import_identifier(m.group("lab_id"))
                    if not ind:
                        self._record_issue(
                            step="step18",
                            sheet="forms_dir",
                            severity="warning",
                            reason="No individual found for form filename.",
                            lab_id=m.group("lab_id"),
                            context={"file": fp.name},
                        )
                        continue
                    if AnalysisRequestForm.objects.filter(file__endswith=fp.name).exists():
                        self._record_issue(
                            step="step18",
                            sheet="forms_dir",
                            severity="info",
                            reason="Form already imported; duplicate skipped.",
                            lab_id=m.group("lab_id"),
                            context={"file": fp.name},
                        )
                        continue
                    if self.dry_run:
                        self.stdout.write(f"  [DRY] form: {fp.name}"); continue
                    with open(fp, "rb") as fh:
                        form_obj = AnalysisRequestForm(
                            individual=ind,
                            description=f"Imported from {fp.name}",
                            created_by=self.admin_user)
                        stored_name = self._fit_uploaded_filename(form_obj.file, fp.name)
                        if stored_name != fp.name:
                            self._record_issue(
                                step="step18",
                                sheet="forms_dir",
                                severity="info",
                                reason="Form filename shortened to fit storage max_length.",
                                lab_id=m.group("lab_id"),
                                context={"original_file": fp.name, "stored_file": stored_name},
                            )
                        form_obj.file.save(stored_name, File(fh))
                        form_obj.save()

        if reports_dir_str:
            reports_dir = Path(reports_dir_str)
            if reports_dir.exists():
                for fp in sorted(reports_dir.glob("*")):
                    if not fp.is_file() or fp.name.startswith("."): continue
                    m = id_regex.match(fp.name)
                    if not m:
                        self._record_issue(
                            step="step18",
                            sheet="reports_dir",
                            severity="warning",
                            reason="Report filename did not match an importable ID pattern.",
                            context={"file": fp.name},
                        )
                        continue
                    ind = find_individual_by_import_identifier(m.group("lab_id"))
                    if not ind:
                        self._record_issue(
                            step="step18",
                            sheet="reports_dir",
                            severity="warning",
                            reason="No individual found for report filename.",
                            lab_id=m.group("lab_id"),
                            context={"file": fp.name},
                        )
                        continue
                    if AnalysisReport.objects.filter(file__endswith=fp.name).exists():
                        self._record_issue(
                            step="step18",
                            sheet="reports_dir",
                            severity="info",
                            reason="Report already imported; duplicate skipped.",
                            lab_id=m.group("lab_id"),
                            context={"file": fp.name},
                        )
                        continue
                    fn_lower = fp.name.lower()
                    pqs = Pipeline.objects.filter(test__sample__individual=ind)
                    target = (
                        pqs.filter(type__name__icontains="wgs").last() if "wgs" in fn_lower
                        else pqs.filter(type__name__icontains="wes").last() if "wes" in fn_lower
                        else pqs.filter(type__name__icontains="sanger").last() if "sanger" in fn_lower
                        else pqs.last())
                    if not target:
                        report_test = ind.get_all_tests().order_by("id").first()
                        if self.dry_run:
                            self._record_issue(
                                step="step18",
                                sheet="reports_dir",
                                severity="info",
                                reason="No matching pipeline found; dry-run would create Franklin fallback pipeline on the individual's first test with Unsure Import.",
                                lab_id=m.group("lab_id"),
                                context={"file": fp.name, "pipeline_type": "Franklin", "test": str(report_test) if report_test else ""},
                            )
                            continue
                        if not report_test:
                            wes_tt = get_or_create_test_type("WES", self.admin_user)
                            self._backfill_testtype_report_fields(wes_tt)
                            sample = ind.samples.first() or self._get_placeholder_sample(ind)
                            report_test = Test.objects.create(
                                sample=sample,
                                test_type=wes_tt,
                                created_by=self.admin_user,
                            )
                            synthetic_statuses = [
                                self.statuses["test"].get("previous"),
                                self.statuses["test"].get("unsure_import"),
                            ]
                            report_test.statuses.set([s for s in synthetic_statuses if s])
                        franklin_type = get_or_create_pipeline_type(
                            "Franklin", self.admin_user, description="Import fallback"
                        )
                        from datetime import date as date_cls
                        target = Pipeline.objects.create(
                            test=report_test,
                            performed_date=date_cls.today(),
                            performed_by=self.admin_user,
                            type=franklin_type,
                            created_by=self.admin_user,
                        )
                        pipeline_unsure_import = self.statuses["pipeline"].get("unsure_import")
                        if pipeline_unsure_import:
                            target.statuses.set([pipeline_unsure_import])
                        self._record_issue(
                            step="step18",
                            sheet="reports_dir",
                            severity="info",
                            reason="No matching pipeline found; created Franklin fallback pipeline on the individual's first test with Unsure Import.",
                            lab_id=m.group("lab_id"),
                            context={"file": fp.name, "pipeline_type": "Franklin", "test": str(report_test)},
                        )
                    if self.dry_run:
                        self.stdout.write(f"  [DRY] report: {fp.name}"); continue
                    with open(fp, "rb") as fh:
                        # Find or create an Analysis on this pipeline
                        target_analysis = target.analyses.first()
                        if not target_analysis:
                            target_analysis = Analysis.objects.create(
                                pipeline=target,
                                created_by=self.admin_user,
                            )
                            target_analysis.performed_by.add(self.admin_user)
                            unsure_import = self.statuses["analysis"].get("unsure_import")
                            if unsure_import:
                                target_analysis.statuses.set([unsure_import])
                        rep = AnalysisReport(
                            analysis=target_analysis,
                            description=f"Imported from {fp.name}",
                            created_by=self.admin_user)
                        stored_name = self._fit_uploaded_filename(rep.file, fp.name)
                        if stored_name != fp.name:
                            self._record_issue(
                                step="step18",
                                sheet="reports_dir",
                                severity="info",
                                reason="Report filename shortened to fit storage max_length.",
                                lab_id=m.group("lab_id"),
                                context={"original_file": fp.name, "stored_file": stored_name},
                            )
                        rep.file.save(stored_name, File(fh))
                        rep.save()
                        report_unsure_import = self.statuses["analysisreport"].get("unsure_import")
                        if report_unsure_import:
                            rep.statuses.set([report_unsure_import])

    # ==================================================================
    # Step 20 — Yayın_İçi
    # ==================================================================

    def _step_yayin_ici(self, yayin_ici_path: str) -> None:
        self.stdout.write(f"Step 20: Yayın_İçi ({yayin_ici_path})…")
        if not os.path.exists(yayin_ici_path):
            self.stdout.write(self.style.ERROR(f"  File not found: {yayin_ici_path}"))
            self._record_issue(
                step="step20",
                sheet="GÜNCELyayıniciyedek",
                severity="error",
                reason="Yayın_İçi workbook file not found.",
                context={"path": yayin_ici_path},
            )
            return

        wb = openpyxl.load_workbook(yayin_ici_path, data_only=True)
        sheet_name = "GÜNCELyayıniciyedek"
        if sheet_name not in wb.sheetnames:
            self.stdout.write(self.style.ERROR(f"  Sheet '{sheet_name}' not found."))
            self._record_issue(
                step="step20",
                sheet=sheet_name,
                severity="error",
                reason="Yayın_İçi sheet not found.",
                context={"path": yayin_ici_path},
            )
            return

        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))
                   if c.value is not None]
        rb_type  = self.id_types.get("RareBoost")
        bb_type  = self.id_types.get("Biobank")
        ct_ind   = ContentType.objects.get_for_model(Individual)
        ct_test  = ContentType.objects.get_for_model(Test)
        updated = skipped = 0
        variant_imported = variant_skipped = 0

        for vals in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in vals): continue
            d = dict(zip(headers, vals[:len(headers)]))
            if not any(
                v is not None and str(v).strip()
                for k, v in d.items()
                if k != "Column 24"
            ):
                continue

            # Lookup
            lab_id = str(d.get("RareBoost ID") or "").strip()
            individual = find_individual_by_rareboost_id(lab_id)
            if not individual:
                bb_val = str(d.get("Biyobanka ID") or "").strip()
                if bb_val and bb_type:
                    individual = Individual.objects.filter(
                        cross_ids__id_type=bb_type, cross_ids__id_value=bb_val).first()
            if not individual:
                self._record_issue(
                    step="step20",
                    sheet=sheet_name,
                    severity="warning",
                    reason="Individual not found by RareBoost ID or Biobank ID.",
                    lab_id=lab_id or d.get("Biyobanka ID"),
                    row=d,
                )
                skipped += 1; continue
            if self.dry_run:
                updated += 1; continue

            changed = False
            # Demographic updates (only if missing)
            sex_val = normalize_sex(d.get("Sex"))
            if sex_val and not individual.sex:
                individual.sex = sex_val; changed = True
            alive_val = to_bool(d.get("Status (ex-alive)"))
            if alive_val is not None and individual.is_alive != alive_val:
                individual.is_alive = alive_val; changed = True
            dob = parse_date(d.get("Date of Birth"))
            if dob and not individual.birth_date:
                individual.birth_date = dob; changed = True
            aoo = str(d.get("Age of Onset") or "").strip()
            if aoo and not individual.age_of_onset:
                individual.age_of_onset = aoo; changed = True
            if changed:
                individual.save()

            # Consanguinity
            consanguinity_raw = d.get("Consanguinity")
            cons = normalize_consanguinity_value(consanguinity_raw)
            if str(consanguinity_raw or "").strip() and individual.family and \
                    individual.family.is_consanguineous != cons:
                individual.family.is_consanguineous = cons
                individual.family.save()

            # Statuses (two separate columns)
            for col in ("Solved (P/LP and clinically relevant VUS), Candidate gene-variant, Unsolved",
                        "NOTE"):
                status_name = str(d.get(col) or "").strip()
                if status_name:
                    st = Status.objects.filter(
                        name=status_name, content_type=ct_ind).first()
                    if st and not individual.statuses.filter(pk=st.pk).exists():
                        individual.statuses.add(st)

            # Disease Group → project
            disease_group = str(d.get("Disease Group") or "").strip()
            if disease_group:
                for pg in [g.strip() for g in disease_group.split(",") if g.strip()]:
                    proj, _ = Project.objects.get_or_create(
                        name=pg,
                        defaults={"created_by": self.admin_user, "priority": "medium"})
                    proj.individuals.add(individual)

            # Institution
            geldi_merkez = str(d.get("Geldiği merkez") or "").strip()
            if geldi_merkez:
                inst, _ = Institution.objects.get_or_create(
                    name=geldi_merkez, defaults={"created_by": self.admin_user})
                individual.institution.add(inst)
                # Physicians for this institution
                clinician_assignments = _build_clinician_assignments(
                    d.get("Klinisyen"),
                    d.get("İletişim Bilgileri - Mail/telefon?"),
                    d.get("İletişim Bilgileri - Telefon/mail?"),
                )
                for klin, contact_values in clinician_assignments:
                    physician = get_or_create_contact(klin, self.admin_user)
                    self._link_physician_to_individual_and_institutions(
                        individual,
                        physician,
                        [inst],
                    )
                    self._apply_contact_details(physician, contact_values)

            # HPO
            hpo_source = d.get("HPO")
            hpo_terms = get_hpo_terms(hpo_source, self.stdout)
            self._apply_hpo_terms_to_individual(individual, hpo_terms, hpo_source)

            # OMIM note
            omim = str(d.get("OMIM") or "").strip()
            if omim:
                parse_and_add_notes(f"OMIM: {omim}", individual, self.admin_user)

            # Previous tests
            prev_raw = d.get("Previous test")
            if prev_raw:
                prev_status = Status.objects.filter(
                    name="Previous", content_type=ct_test).first()
                for tt_name in self._normalize_test_tokens(str(prev_raw)):
                    tt = get_or_create_test_type(tt_name, self.admin_user)
                    self._backfill_testtype_report_fields(tt)
                    if not individual.get_all_tests().filter(test_type=tt).exists():
                        sample = (individual.samples.first()
                                  or self._get_placeholder_sample(individual))
                        test = Test.objects.create(
                            sample=sample, test_type=tt, created_by=self.admin_user)
                        synthetic_statuses = [
                            prev_status,
                            self.statuses["test"].get("unsure_import"),
                        ]
                        test.statuses.set([s for s in synthetic_statuses if s])

            # RareBoost Reanaliz/WGS/WES/RNA seq
            rb_raw = d.get("RareBoost Reanaliz/WGS/WES/RNA seq")
            if rb_raw:
                self._process_rb_reanaliz(individual, str(rb_raw))

            # Singleton-Trio notes on non-previous tests
            singleton_trio = str(d.get("Singleton-Trio") or "").strip()
            if singleton_trio:
                non_prev_tests = list(
                    individual.get_all_tests()
                    .exclude(statuses__name="Previous")
                    .order_by("id"))
                trio_values = [v.strip() for v in singleton_trio.split(",") if v.strip()]
                if not trio_values:
                    trio_values = [singleton_trio]
                last_val = trio_values[-1]
                for idx, test_obj in enumerate(non_prev_tests):
                    note_val = trio_values[idx] if idx < len(trio_values) else last_val
                    parse_and_add_notes(note_val, test_obj, self.admin_user)

            # Variant import — prefer the importable Chromosomal Position column.
            variant_text = str(
                d.get("Chromosomal Position")
                or d.get("Variant")
                or ""
            ).strip()
            if variant_text:
                for line in _split_yayin_variant_text(variant_text):
                    records = _extract_variant_records(line)
                    if not records:
                        variant_skipped += 1
                        self._record_issue(
                            step="step20",
                            sheet=sheet_name,
                            severity="warning",
                            reason="Variant line could not be imported safely; no genomic coordinates recognized.",
                            lab_id=lab_id,
                            row=d,
                            context={"variant_line": line},
                        )
                        self.stdout.write(self.style.WARNING(
                            f"  Yayın_İçi: skipped variant for {lab_id} -> {line}"))
                        continue

                    for record in records:
                        zyg = _normalize_yayin_zygosity(d.get("Zygosity"))
                        model_kind = record["kind"]
                        model_cls = {
                            "snv": SNV,
                            "delins": delins,
                            "cnv": CNV,
                            "sv": SV,
                        }.get(model_kind)
                        if not model_cls:
                            variant_skipped += 1
                            self._record_issue(
                                step="step20",
                                sheet=sheet_name,
                                severity="warning",
                                reason="Variant record had an unsupported model kind.",
                                lab_id=lab_id,
                                row=d,
                                context={"variant_line": line, "record": record},
                            )
                            continue

                        try:
                            if model_cls in (SNV, delins):
                                lookup = {
                                    "individual": individual,
                                    "chromosome": record["chromosome"],
                                    "start": record["start"],
                                    "reference": record["reference"],
                                    "alternate": record["alternate"],
                                }
                                defaults = {
                                    "end": record["end"],
                                    "zygosity": zyg,
                                    "created_by": self.admin_user,
                                }
                            elif model_cls is CNV:
                                lookup = {
                                    "individual": individual,
                                    "chromosome": record["chromosome"],
                                    "start": record["start"],
                                    "end": record["end"],
                                    "cnv_type": record.get("cnv_type", "gain"),
                                    "copy_number": record.get("copy_number"),
                                }
                                defaults = {
                                    "zygosity": zyg,
                                    "created_by": self.admin_user,
                                }
                            else:  # SV
                                lookup = {
                                    "individual": individual,
                                    "chromosome": record["chromosome"],
                                    "start": record["start"],
                                    "end": record["end"],
                                    "sv_type": record.get("sv_type", "deletion"),
                                }
                                defaults = {
                                    "zygosity": zyg,
                                    "created_by": self.admin_user,
                            }

                            variant_obj, created_variant = model_cls.objects.get_or_create(
                                **lookup, defaults=defaults)

                            if not created_variant:
                                variant_skipped += 1
                                self._record_issue(
                                    step="step20",
                                    sheet=sheet_name,
                                    severity="info",
                                    reason="Variant already exists; duplicate skipped.",
                                    lab_id=lab_id,
                                    row=d,
                                    context={"variant_line": line, "record": record, "source_column": "Chromosomal Position"},
                                )
                                continue

                            if record.get("note"):
                                parse_and_add_notes(record["note"], variant_obj, self.admin_user)

                            variant_imported += 1
                        except Exception as exc:
                            variant_skipped += 1
                            self._record_issue(
                                step="step20",
                                sheet=sheet_name,
                                severity="error",
                                reason="Failed to import parsed variant record.",
                                lab_id=lab_id,
                                row=d,
                                context={"variant_line": line, "record": record, "error": str(exc)},
                            )
                            self.stdout.write(self.style.ERROR(
                                f"  Yayın_İçi: variant import failed for {lab_id} -> {line}: {exc}"))

            updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"  Yayın_İçi: updated={updated} skipped={skipped} variants_imported={variant_imported} variants_skipped={variant_skipped}"))

    def _normalize_test_tokens(self, text: str) -> list:
        """Split and normalise test type strings from the Previous test column."""
        text = text.replace('"Gene Panel, Single Gene"', "Targeted Panel")
        tokens = [t.strip() for t in text.replace("\n", ",").split(",") if t.strip()]
        result = []
        skip_next = False
        for i, tok in enumerate(tokens):
            if skip_next:
                skip_next = False; continue
            tl = tok.lower()
            if tl == "gene panel" and i + 1 < len(tokens) \
                    and tokens[i + 1].lower() == "single gene":
                result.append("Targeted Panel"); skip_next = True
            elif tl in {"wes", "wgs", "cma", "karyotype"}:
                result.append(tok.upper())
            elif tl in {"targeted panel", "gene panel single gene"}:
                result.append("Targeted Panel")
            elif tl in {"rna seq", "rnaseq", "rna-seq"}:
                result.append("RNA Seq")
            else:
                result.append(tok.strip('" '))
        return [r for r in result if r]

    def _process_rb_reanaliz(self, individual, text: str) -> None:
        """Handle the 'RareBoost Reanaliz/WGS/WES/RNA seq' column."""
        parts = [p.strip() for p in text.replace("\n", ",").split(",") if p.strip()]
        if not parts:
            return

        def _normalize_analysis_token(part: str) -> str:
            lower = part.lower().strip()
            if "rna seq" in lower or "rnaseq" in lower or "rna-seq" in lower:
                return "RNA Seq"
            if "targeted panel" in lower:
                return "Targeted Panel"
            if "wes" in lower:
                return "WES"
            if "wgs" in lower:
                return "WGS"
            return part.strip()

        normalized_tokens: list[str] = []
        seen_tokens: set[str] = set()
        for part in parts:
            token = _normalize_analysis_token(part)
            if token and token not in seen_tokens:
                normalized_tokens.append(token)
                seen_tokens.add(token)

        if not normalized_tokens:
            return

        unsure_import = self.statuses["analysis"].get("unsure_import")
        analysis_type_cache: dict[str, object] = {}
        from datetime import date as date_cls

        for token in normalized_tokens:
            analysis_type = analysis_type_cache.get(token)
            if analysis_type is None:
                analysis_type = get_or_create_analysis_type(token, self.admin_user)
                analysis_type_cache[token] = analysis_type

            target_pipeline = (
                Pipeline.objects.filter(
                    test__sample__individual=individual,
                    type__name__iexact=token,
                )
                .order_by("id")
                .first()
            )
            if not target_pipeline:
                target_pipeline = (
                    Pipeline.objects.filter(
                        test__sample__individual=individual,
                        type__name__icontains=token,
                    )
                    .order_by("id")
                    .first()
                )

            analysis = Analysis.objects.create(
                pipeline=target_pipeline,
                type=analysis_type,
                performed_date=date_cls.today(),
                created_by=self.admin_user,
            )
            analysis.performed_by.add(self.admin_user)
            parse_and_add_notes(
                "\n".join(filter(None, [
                    f"RareBoost Reanaliz/WGS/WES/RNA seq: {text}",
                    f"Normalized token: {token}",
                ])),
                analysis,
                self.admin_user,
            )

            if not target_pipeline:
                self.stdout.write(self.style.WARNING(
                    f"  RareBoost reanalysis: no matching pipeline found for {individual.primary_id} -> {token}; created Analysis with Unsure Import"))
                self._record_issue(
                    step="step20",
                    sheet="GÜNCELyayıniciyedek",
                    severity="warning",
                    reason="No matching pipeline found for RareBoost reanalysis entry; created Analysis with Unsure Import.",
                    lab_id=individual.primary_id,
                    context={
                        "raw_value": text,
                        "normalized_token": token,
                    },
                )
                if unsure_import:
                    analysis.statuses.set([unsure_import])

    # ==================================================================
    # Internal helpers
    # ==================================================================

    def _get_placeholder_sample(self, individual) -> Sample:
        existing = individual.samples.first()
        if existing:
            return existing
        placeholder_type = get_or_create_sample_type("Placeholder", self.admin_user)
        placeholder_contact = get_or_create_contact_for_user(self.admin_user, self.admin_user)
        sample = Sample.objects.create(
            individual=individual,
            sample_type=placeholder_type,
            isolation_by=placeholder_contact,
            created_by=self.admin_user)
        sample_statuses = [
            self.statuses["sample"].get("not_available"),
            self.statuses["sample"].get("unsure_import"),
        ]
        sample.statuses.set([s for s in sample_statuses if s])
        return sample
