import openpyxl
import json
import os
import csv
from datetime import datetime
from django.core.management.base import BaseCommand
from django.core.management import call_command
from django.contrib.auth.models import User
from django.contrib.contenttypes.models import ContentType
from django.utils import timezone
from lab.models import (
    Family,
    Individual,
    Sample,
    Test,
    Status,
    SampleType,
    Institution,
    TestType,
    Analysis,
    AnalysisType,
    Note,
    IdentifierType,
    CrossIdentifier,
    Project,
    Task
)
from variant.models import Variant, SNV, Gene
from ontologies.models import Term, Ontology
import re
from django.contrib.auth import get_user_model

User = get_user_model()


class Command(BaseCommand):
    help = 'Import data from OZBEK lab google sheets file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the TSV file')
        parser.add_argument('--admin-username', type=str, help='Admin username for created_by fields')

    def _parse_date(self, date_str):
        """Parse date string in various formats"""
        if not date_str:
            return None

        # Handle datetime objects from Excel
        if hasattr(date_str, 'date'):
            return date_str.date()

        # Handle string dates
        if isinstance(date_str, str):
            formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']
            for fmt in formats:
                try:
                    return datetime.strptime(date_str.strip(), fmt).date()
                except ValueError:
                    continue
        return None

    def _get_family_id(self, lab_id):
        """Extract family ID from lab_id"""
        if not lab_id:
            return None
        return lab_id.split('.')[0]

    def _get_or_create_user(self, name, admin_user):
        """Get or create a user based on name"""
        if not name:
            return admin_user
            
        # Try to find user by name
        user = User.objects.filter(username__icontains=name).first()
        if user:
            return user
            
        # Create new user if not found
        username = name.lower().replace(' ', '_')
        user = User.objects.create_user(
            username=username,
            email=f'{username}@example.com',
            password='changeme123',
            first_name=name.split()[0] if ' ' in name else name,
            last_name=name.split()[1] if ' ' in name else ''
        )
        return user

    def _get_or_create_institution(self, name, contact, admin_user):
        """Get or create an institution"""
        if not name:
            return None
            
        institution = Institution.objects.filter(name=name).first()
        if institution:
            return institution
            
        institution = Institution.objects.create(
            name=name,
            contact=contact,
            created_by=admin_user
        )
        return institution

    def _get_or_create_sample_type(self, name, admin_user):
        """Get or create a sample type"""
        if not name:
            return None
            
        sample_type = SampleType.objects.filter(name=name).first()
        if sample_type:
            return sample_type
            
        sample_type = SampleType.objects.create(
            name=name,
            created_by=admin_user
        )
        return sample_type

    def _get_hpo_terms(self, hpo_codes_str):
        """Get HPO terms from newline-separated descriptions and codes"""
        if not hpo_codes_str:
            return []
            
        # Get the HP ontology
        hp_ontology = Ontology.objects.filter(type=1).first()  # 1 is HP in ONTOLOGY_CHOICES
        if not hp_ontology:
            self.stdout.write(self.style.WARNING('HP ontology not found'))
            return []
            
        # Split into individual terms (each term is "Description HP:code")
        terms = []
        
        # Split by newlines and filter out empty lines and '+' lines
        term_strings = []
        for line in hpo_codes_str.split('\n'):
            line = line.strip()
            if line and line != '+':
                term_strings.append(line)
        
        for term_str in term_strings:
            # Find the HP: code in the term string
            hp_match = re.search(r'HP:(\d+)', term_str)
            if not hp_match:
                self.stdout.write(self.style.WARNING(f'No HP code found in term: {term_str}'))
                continue
                
            code = hp_match.group(1)
            # Get the description (everything before the HP: code)
            description = term_str[:hp_match.start()].strip()
            
            # Find the term by identifier
            term = Term.objects.filter(
                ontology=hp_ontology,
                identifier=code
            ).first()
            
            if term:
                # self.stdout.write(self.style.SUCCESS(f'Found HPO term: {term.label} (HP:{code})'))
                terms.append(term)
            else:
                self.stdout.write(self.style.WARNING(f'HPO term not found: {description} (HP:{code})'))
                # Try to find by label as fallback
                term = Term.objects.filter(
                    ontology=hp_ontology,
                    label__icontains=description
                ).first()
                if term:
                    self.stdout.write(self.style.SUCCESS(f'Found HPO term by label: {term.label} (HP:{term.identifier})'))
                    terms.append(term)
        
        self.stdout.write(self.style.SUCCESS(f'Total HPO terms found: {len(terms)}'))
        return terms



    def _parse_and_add_notes(self, note_text, target_obj, admin_user):
        """Parse newline-separated notes and add as separate Note objects.

        - Each non-empty line becomes a separate Note with full line as content
        - If a line begins with a date in format DD.MM.YYYY, set the earliest
          history record's history_date for that note to this date (midnight)
        """
        if not note_text:
            return
        # Ensure we have a content type for the target object
        try:
            ct = ContentType.objects.get_for_model(target_obj)
        except Exception:
            return

        lines = str(note_text).splitlines()
        for raw_line in lines:
            line = (raw_line or '').strip()
            if not line:
                continue
            # Create the note with the full line content
            try:
                note = Note.objects.create(
                    content=line,
                    user=admin_user,
                    content_type=ct,
                    object_id=getattr(target_obj, 'pk', None) or getattr(target_obj, 'id', None),
                )
            except Exception:
                continue

            # If line starts with a date, set the note's earliest history_date
            m = re.match(r'^(\d{2}\.\d{2}\.\d{4})', line)
            if m:
                try:
                    dt = datetime.strptime(m.group(1), '%d.%m.%Y')
                    # Ensure timezone-aware datetime to avoid warnings with USE_TZ
                    if timezone.is_naive(dt):
                        dt = timezone.make_aware(dt, timezone.get_current_timezone())
                    first_hist = note.history.earliest()
                    if first_hist:
                        first_hist.history_date = dt
                        first_hist.save()
                except Exception:
                    # If anything goes wrong, skip adjusting history date
                    pass

    def analysis_add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the TSV file')
        parser.add_argument(
            '--admin-user',
            type=str,
            help='Username of the admin user to use for created_by fields',
            required=True
        )

    def __analysis_parse_date(self, date_str):
        if not date_str or date_str.lower() == 'na':
            return None
        try:
            return datetime.strptime(date_str, '%d.%m.%Y').date()
        except ValueError:
            return None

    def _get_or_create_status(self, name, description, color, admin_user, content_type=None, icon=None):
        if not name:
            return None
        status, created = Status.objects.get_or_create(
            name=name,
            content_type=content_type,
            defaults={
                'description': description or '',
                'color': color or '#000000',
                'created_by': admin_user,
                'icon': icon,
            }
        )
        # Backfill or update icon if needed
        if icon and (created or not status.icon or status.icon != icon):
            status.icon = icon
            status.save()
        return status

    def _get_or_create_test_type(self, name, admin_user):
        if not name:
            return None
        test_type, created = TestType.objects.get_or_create(
            name=name,
            defaults={'created_by': admin_user}
        )
        return test_type

    def _get_or_create_analysis_type(self, name, description, admin_user):
        if not name:
            return None
        analysis_type, created = AnalysisType.objects.get_or_create(
            name=name,
            defaults={
                'description': description or '',
                'created_by': admin_user
            }
        )
        return analysis_type

    def _get_or_create_placeholder_sample(self, individual, admin_user):
        # Create a placeholder sample type if it doesn't exist
        placeholder_type = self._get_or_create_sample_type('Placeholder', admin_user)
        
        # Create a placeholder sample
        sample = Sample.objects.create(
            individual=individual,
            sample_type=placeholder_type,
            status=Status.objects.get(name='Not Available', content_type=ContentType.objects.get(app_label='lab', model='sample')),
            isolation_by=admin_user,
            created_by=admin_user
        )
        return sample

    def _get_initials(self, full_name: str) -> str:
        parts = [p for p in (full_name or '').split() if p]
        return ''.join(part[0].upper() for part in parts)

    def _parse_variant_string(self, variant_str):
        """
        Parse variant string like 'chr10-77984023 A>G'
        Returns (chromosome, start, reference, alternate)
        """
        if not variant_str:
            return None
        
        # Regex for chrX-12345 REF>ALT
        match = re.match(r'(chr[\w]+)-(\d+)\s+([ACGT]+)>([ACGT]+)', variant_str.strip())
        if match:
            return match.groups()
        return None

    def _map_zygosity(self, zygosity_str):
        if not zygosity_str:
            return 'unknown'
        
        z_lower = zygosity_str.lower().strip()
        if 'het' in z_lower:
            return 'het'
        if 'hom' in z_lower:
            return 'hom'
        if 'hemi' in z_lower:
            return 'hemi'
        return 'unknown'

    def handle(self, *args, **options):

        call_command('create_ozbek_users') # Create groups and Ozbek users

        file_path = options['file_path']
        ozbek_lab_sheet = 'OZBEK LAB'
        analiz_takip_sheet = 'Analiz Takip'
        admin_username = options.get('admin_username')
        if not admin_username:
            self.stdout.write(self.style.ERROR('Please provide admin username with --admin-username'))
            return
        try:
            admin_user = User.objects.get(username=admin_username)
        except User.DoesNotExist:
            self.stdout.write(self.style.ERROR(f'Admin user {admin_username} not found'))
            return

        # --- CREATE ALL STATUSES FIRST ---
        self.stdout.write('Creating all required statuses...')
        
        # Individual statuses
        individual_statuses = {
            'Registered': {'description': 'Initial status for new entries', 'color': 'gray', 'icon': 'fa-user-plus'},
            'Solved': {'description': 'Entry has been completed', 'color': 'green', 'icon': 'fa-circle-check'}
        }
        for status_name, status_data in individual_statuses.items():
            status, created = Status.objects.get_or_create(
                name=status_name,
                content_type=ContentType.objects.get(app_label='lab', model='individual'),
                defaults={
                    'description': status_data['description'],
                    'color': status_data['color'],
                    'created_by': admin_user,
                    'icon': status_data.get('icon'),
                }
            )
            # Backfill icon if missing or different
            desired_icon = status_data.get('icon')
            if desired_icon and (created or not status.icon or status.icon != desired_icon):
                status.icon = desired_icon
                status.save()
            if created:
                self.stdout.write(self.style.SUCCESS(f'Created individual status: {status_name}'))
            else:
                self.stdout.write(f'Individual status ensured: {status_name}')

        # Sample statuses
        sample_statuses = {
            'placeholder': self._get_or_create_status(
                'Not Available',
                'A placeholder sample for tests performed off-center',
                'gray',
                admin_user,
                ContentType.objects.get(app_label='lab', model='sample'),
                icon='fa-ban'
            ),
            'pending_blood_recovery': self._get_or_create_status(
                'Pending Blood Recovery',
                'Awaiting blood draw',
                'red',
                admin_user,
                ContentType.objects.get(app_label='lab', model='sample'),
                icon='fa-droplet'
            ),
            'pending_isolation': self._get_or_create_status(
                'Pending Isolation',
                'Awaiting isolation of sample',
                'yellow',
                admin_user,
                ContentType.objects.get(app_label='lab', model='sample'),
                icon='fa-vials'
            ),
            'available': self._get_or_create_status(
                'Available',
                'Available for tests',
                'green',
                admin_user,
                ContentType.objects.get(app_label='lab', model='sample'),
                icon='fa-circle-check'
            ),
        }

        # Project statuses
        imported_project_status = self._get_or_create_status(
            'In Progress',
            'Project is in progress',
            'green',
            admin_user,
            ContentType.objects.get_for_model(Project),
            icon='fa-diagram-project'
        )
        
        self._get_or_create_status(
            'Setting Up',
            'Project is being set up',
            'yellow',
            admin_user,
            ContentType.objects.get_for_model(Project),
            icon='fa-gears'
        )
        
        self._get_or_create_status(
            'Completed',
            'Project is completed',
            'gray',
            admin_user,
            ContentType.objects.get_for_model(Project),
            icon='fa-flag-checkered'
        )

        # Analysis statuses
        completed_status = self._get_or_create_status(
            'Completed',
            'Analysis completed',
            'green',
            admin_user,
            ContentType.objects.get_for_model(Analysis),
            icon='fa-circle-check'
        )

        self._get_or_create_status(
            'In Progress',
            'Analysis is in progress',
            'yellow',
            admin_user,
            ContentType.objects.get_for_model(Analysis),
            icon='fa-spinner'
        )

        self._get_or_create_status(
            'Pending Data',
            'Analysis is pending data',
            'red',
            admin_user,
            ContentType.objects.get_for_model(Analysis),
            icon='fa-hourglass-half'
        )

        # Test statuses
        self._get_or_create_status(
            'Completed',
            'Test completed',
            'green',
            admin_user,
            ContentType.objects.get_for_model(Test),
            icon='fa-circle-check'
        )

        self._get_or_create_status(
            'In Progress',
            'Test is in progress',
            'yellow',
            admin_user,
            ContentType.objects.get_for_model(Test),
            icon='fa-spinner'
        )

        self._get_or_create_status(
            'Pending',
            'Test is pending',
            'red',
            admin_user,
            ContentType.objects.get_for_model(Test),
            icon='fa-clock'
        )

        # Task statuses
        self._get_or_create_status(
            'Active',
            'Task is ongoing',
            'yellow',
            admin_user,
            ContentType.objects.get_for_model(Task),
            icon='fa-list-check'
        )

        self._get_or_create_status(
            'Completed',
            'Task is completed',
            'green',
            admin_user,
            ContentType.objects.get_for_model(Task),
            icon='fa-circle-check'
        )

        self._get_or_create_status(
            'Overdue',
            'Task is overdue',
            'red',
            admin_user,
            ContentType.objects.get_for_model(Task),
            icon='fa-triangle-exclamation'
        )

        # --- CREATE IdentifierType objects for cross IDs ---
        id_types = {}
        for idtype_name in ["RareBoost", "Biobank", "ERDERA"]:
            id_type, _ = IdentifierType.objects.get_or_create(
                name=idtype_name,
                defaults={
                    "description": f"{idtype_name} identifier type",
                    "created_by": admin_user
                }
            )
            id_types[idtype_name] = id_type

        # Create Unknown institution if it doesn't exist
        self.stdout.write('Ensuring Unknown institution exists...')
        unknown_institution, created = Institution.objects.get_or_create(
            name='Unknown',
            defaults={
                'contact': 'Unknown institution - placeholder for missing institution data',
                'created_by': admin_user
            }
        )
        if created:
            self.stdout.write(self.style.SUCCESS('Created Unknown institution'))
        else:
            self.stdout.write('Unknown institution already exists')

        # Note: Projects will now be created per-individual based on the new
        # 'Projeler' column in the OZBEK LAB sheet.

        # --- XLSX Reading ---
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # --- INSTITUTION MAP SHEET (Coordinates) ---
        coords_map = {}
        try:
            ws_map = wb['Gönderen Kurum Harita']
            headers_map = [cell.value for cell in next(ws_map.iter_rows(min_row=1, max_row=1))]
            headers_map = [h for h in headers_map if h is not None]
            name_key = 'Kurum'
            coord_key = 'Harita'
            city_key = 'Şehir'
            official_name_key = 'Resmi Ad'
            total_rows = 0
            empty_rows = 0
            parsed_rows = 0
            skipped_no_name_or_coord = 0
            parse_errors = 0
            for row in ws_map.iter_rows(min_row=2, values_only=True):
                total_rows += 1
                if all(cell is None for cell in row):
                    empty_rows += 1
                    continue
                row = row[:len(headers_map)]
                row_dict = dict(zip(headers_map, row))
                raw_name = row_dict.get(name_key)
                coord_val = row_dict.get(coord_key)
                city_val = row_dict.get(city_key)
                official_name_val = row_dict.get(official_name_key)
                if not raw_name or not coord_val:
                    skipped_no_name_or_coord += 1
                    continue
                # Normalize and support multiple comma-separated names similar to source sheet
                for name in [n.strip() for n in str(raw_name).split(',') if n and str(n).strip()]:
                    try:
                        lat_str, lon_str = [p.strip() for p in str(coord_val).split(',')[:2]]
                        lat_val = float(lat_str)
                        lon_val = float(lon_str)
                        if name not in coords_map:
                            coords_map[name] = {}
                        coords_map[name]['coords'] = (lat_val, lon_val)
                        parsed_rows += 1
                    except Exception:
                        parse_errors += 1
                        continue
                    if city_val:
                        coords_map[name]['city'] = city_val.strip()
                    if official_name_val:
                        coords_map[name]['official_name'] = official_name_val.strip()
            self._institution_info = coords_map
        except KeyError:
            # Sheet not present; proceed without coordinates
            self._institution_info = {}
            print("[Map Import] ERROR: coordinates sheet 'Gönderen Kurum Harita' not found; proceeding without coordinates")
        # --- OZBEK LAB SHEET ---
        ws_lab = wb[ozbek_lab_sheet]
        headers_lab = [cell.value for cell in next(ws_lab.iter_rows(min_row=1, max_row=1))]
        # Filter out None column names (empty columns at the end)
        headers_lab = [h for h in headers_lab if h is not None]
        
        # First pass: Collect unique values
        unique_family_ids = set()
        institution_details = {}
        for row in ws_lab.iter_rows(min_row=2, values_only=True):
            # Skip rows where all values are None (empty rows at the end)
            if all(cell is None for cell in row):
                continue
            # Truncate row to match filtered headers (remove empty columns)
            row = row[:len(headers_lab)]
            row_dict = dict(zip(headers_lab, row))
            lab_id = row_dict.get('Özbek Lab. ID')
            if not lab_id:
                self.stdout.write(f'Skipping row - missing lab ID. Row data: {row_dict}')
                continue
            if lab_id:
                family_id = self._get_family_id(lab_id)
                if family_id:
                    unique_family_ids.add(family_id)
            institution_name_raw = row_dict.get('Gönderen Kurum/Birim')
            contact_info = row_dict.get('Klinisyen & İletişim Bilgileri')
            if institution_name_raw:
                institution_names = [n.strip() for n in str(institution_name_raw).split(',') if n and str(n).strip()]
                for institution_name in institution_names:
                    if institution_name not in institution_details:
                        institution_details[institution_name] = set()
                    if contact_info:
                        institution_details[institution_name].add(contact_info)
        # Create Families
        families = {}
        for family_id in unique_family_ids:
            family, _ = Family.objects.get_or_create(
                family_id=family_id,
                defaults={'created_by': admin_user}
            )
            families[family_id] = family
        # Create Institutions with contact information (and coordinates if available)
        institutions = {}
        for name, contacts in institution_details.items():
            institution, _ = Institution.objects.get_or_create(
                name=name,
                defaults={
                    'contact': '\n'.join(contacts) if contacts else '',
                    'created_by': admin_user,
                    'latitude': self._institution_info.get(name, {}).get('coords', (None, None))[0] if self._institution_info.get(name) else 0.0,
                    'longitude': self._institution_info.get(name, {}).get('coords', (None, None))[1] if self._institution_info.get(name) else 0.0,
                    'city': self._institution_info.get(name, {}).get('city', ''),
                    'official_name': self._institution_info.get(name, {}).get('official_name', ''),
                }
            )
            if not _ and contacts:
                existing_contacts = set(institution.contact.split('\n')) if institution.contact else set()
                all_contacts = existing_contacts.union(contacts)
                institution.contact = '\n'.join(all_contacts)
                # Backfill coordinates if available and not set (or set to defaults)
                if name in self._institution_coords:
                    lat_val, lon_val = self._institution_coords[name]
                    if (institution.latitude in (None, 0.0)) and lat_val is not None:
                        institution.latitude = lat_val
                    if (institution.longitude in (None, 0.0)) and lon_val is not None:
                        institution.longitude = lon_val
                institution.save()
            institutions[name] = institution
        # Second pass: Create Individuals and CrossIdentifiers
        for row in ws_lab.iter_rows(min_row=2, values_only=True):
            # Skip rows where all values are None (empty rows at the end)
            if all(cell is None for cell in row):
                continue
            # Truncate row to match filtered headers (remove empty columns)
            row = row[:len(headers_lab)]
            row_dict = dict(zip(headers_lab, row))
            try:
                lab_id = row_dict.get('Özbek Lab. ID')
                if not lab_id:
                    self.stdout.write(f'Skipping row - missing lab ID. Row data: {row_dict}')
                    continue
                family_id = self._get_family_id(lab_id)
                if not family_id or family_id not in families:
                    self.stdout.write(f'Skipping row - invalid family_id: {family_id} for lab_id: {lab_id}')
                    continue
                institution_name_raw = row_dict.get('Gönderen Kurum/Birim')
                institution_list = []
                if institution_name_raw:
                    for inst_name in [n.strip() for n in str(institution_name_raw).split(',') if n and str(n).strip()]:
                        institution_list.append(institutions.get(inst_name, unknown_institution))
                if not institution_list:
                    institution_list = [unknown_institution]
                is_index = False
                if re.search(r'\.1(\.|$)', lab_id):
                    is_index = True
                full_name = row_dict.get('Ad-Soyad')
                if not full_name:
                    self.stdout.write(f'Skipping row - missing full_name for lab_id: {lab_id}')
                    continue
                individual = Individual.objects.filter(full_name=full_name, family=families[family_id]).first()
                if not individual:
                    tc_identity_val = row_dict.get('TC Kimlik No')
                    if tc_identity_val is not None:
                        if isinstance(tc_identity_val, str) and tc_identity_val.strip() == '':
                            tc_identity_val = None
                        elif isinstance(tc_identity_val, (int, float)):
                            tc_identity_val = int(tc_identity_val)
                        else:
                            try:
                                tc_identity_val = int(tc_identity_val)
                            except (TypeError, ValueError):
                                tc_identity_val = None
                    individual = Individual.objects.create(
                        full_name=full_name,
                        family=families[family_id],
                        birth_date=self._parse_date(row_dict.get('Doğum Tarihi')),
                        icd11_code='',
                        status=Status.objects.get(name='Registered', content_type=ContentType.objects.get(app_label='lab', model='individual')),
                        created_by=admin_user,
                        diagnosis='',
                        diagnosis_date=None,
                        council_date=None,
                        is_index=is_index,
                        tc_identity=tc_identity_val
                    )
                    if institution_list:
                        individual.institution.set(institution_list)
                    self.stdout.write(self.style.SUCCESS(f"Created individual: {self._get_initials(full_name)}"))
                    # Add individual to projects listed in 'Projeler'
                    projects_field = row_dict.get('Projeler')
                    if projects_field:
                        project_names = [p.strip() for p in str(projects_field).split(',') if p and str(p).strip()]
                        for pname in project_names:
                            project_obj, _ = Project.objects.get_or_create(
                                name=pname,
                                defaults={
                                    'description': '',
                                    'created_by': admin_user,
                                    'status': imported_project_status,
                                    'priority': 'medium'
                                }
                            )
                            project_obj.individuals.add(individual)
                else:
                    if individual.is_index != is_index:
                        individual.is_index = is_index
                        individual.save()
                    self.stdout.write(f'Individual already exists: {self._get_initials(full_name)}')
                # Add individual-level notes (Kurum Notları)
                kurum_notes = row_dict.get('Kurum Notları')
                if kurum_notes:
                    self._parse_and_add_notes(kurum_notes, individual, admin_user)
                # Add individual-level follow-up notes (Takip Notları)
                takip_notes = row_dict.get('Takip Notları')
                if takip_notes:
                    self._parse_and_add_notes(takip_notes, individual, admin_user)
                # CrossIdentifiers
                rareboost_id = row_dict.get('Özbek Lab. ID')
                if rareboost_id:
                    CrossIdentifier.objects.get_or_create(
                        individual=individual,
                        id_type=id_types["RareBoost"],
                        defaults={
                            'id_value': rareboost_id,
                            'created_by': admin_user
                        }
                    )
                biobank_id = row_dict.get('Biyobanka ID')
                if biobank_id:
                    CrossIdentifier.objects.get_or_create(
                        individual=individual,
                        id_type=id_types["Biobank"],
                        defaults={
                            'id_value': biobank_id,
                            'created_by': admin_user
                        }
                    )
                erdera_id = row_dict.get('ERDERA ID')
                if erdera_id:
                    CrossIdentifier.objects.get_or_create(
                        individual=individual,
                        id_type=id_types["ERDERA"],
                        defaults={
                            'id_value': erdera_id,
                            'created_by': admin_user
                        }
                    )
            except Exception as e:
                self.stdout.write(
                    self.style.ERROR(f"Error processing row for {row_dict.get('Özbek Lab. ID')}: {str(e)}")
                )
                continue
        # --- END: Individual and CrossIdentifier logic ---
        # Create leftovers file in the same directory as input file
        leftovers_path = os.path.join(os.path.dirname(file_path), 'import_ozbek_lab_leftovers.tsv')
        leftover_rows = []
        for row in ws_lab.iter_rows(min_row=2, values_only=True):
            # Skip rows where all values are None (empty rows at the end)
            if all(cell is None for cell in row):
                continue
            # Truncate row to match filtered headers (remove empty columns)
            row = row[:len(headers_lab)]
            row_dict = dict(zip(headers_lab, row))
            lab_id = row_dict.get('Özbek Lab. ID')
            if not lab_id:
                self.stdout.write(f'Skipping row - missing lab ID. Row data: {row_dict}')
                leftover_rows.append(row_dict)
                continue
            try:
                individual = Individual.objects.filter(cross_ids__id_value=lab_id).first()
                if not individual:
                    self.stdout.write(f'Individual not found with lab_id: {lab_id}. Available cross_ids: {list(Individual.objects.values_list("cross_ids__id_value", flat=True))[:10]}')
                    leftover_rows.append(row_dict)
                    continue
                self.stdout.write(f'Found existing individual: {lab_id}')
            except Exception as e:
                leftover_rows.append(row_dict)
                self.stdout.write(self.style.WARNING(f'Error finding individual for {lab_id}: {str(e)}'))
                continue
            institution_name_raw = row_dict.get('Gönderen Kurum/Birim')
            institution_objs = []
            if institution_name_raw:
                for name in [n.strip() for n in str(institution_name_raw).split(',') if n and str(n).strip()]:
                    inst_obj = self._get_or_create_institution(
                        name,
                        row_dict.get('Klinisyen & İletişim Bilgileri'),
                        admin_user
                    )
                    if inst_obj:
                        institution_objs.append(inst_obj)
            if not institution_objs:
                institution_objs = [unknown_institution]
            is_index = False
            if re.search(r'\.1(\.|$)', lab_id):
                is_index = True
            individual.full_name = row_dict.get('Ad-Soyad', individual.full_name)
            tc_identity_val = row_dict.get('TC Kimlik No', individual.tc_identity)
            if tc_identity_val is not None:
                if isinstance(tc_identity_val, str) and tc_identity_val.strip() == '':
                    tc_identity_val = None
                elif isinstance(tc_identity_val, (int, float)):
                    tc_identity_val = int(tc_identity_val)
                else:
                    try:
                        tc_identity_val = int(tc_identity_val)
                    except (TypeError, ValueError):
                        tc_identity_val = None
            individual.tc_identity = tc_identity_val
            individual.birth_date = self._parse_date(row_dict.get('Doğum Tarihi')) or individual.birth_date
            individual.icd11_code = row_dict.get('ICD11', individual.icd11_code)
            if institution_objs:
                individual.institution.set(institution_objs)
            individual.status = Status.objects.get(name='Registered', content_type=ContentType.objects.get(app_label='lab', model='individual'))
            individual.is_index = is_index
            individual.save()
            # Add/ensure individual-project associations from 'Projeler'
            projects_field = row_dict.get('Projeler')
            if projects_field:
                project_names = [p.strip() for p in str(projects_field).split(',') if p and str(p).strip()]
                for pname in project_names:
                    project_obj, _ = Project.objects.get_or_create(
                        name=pname,
                        defaults={
                            'description': '',
                            'created_by': admin_user,
                            'status': imported_project_status,
                            'priority': 'medium'
                        }
                    )
                    project_obj.individuals.add(individual)
            hpo_terms = self._get_hpo_terms(row_dict.get('HPO kodları'))
            if hpo_terms:
                individual.hpo_terms.add(*hpo_terms)
                self.stdout.write(self.style.SUCCESS(f'Added {len(hpo_terms)} HPO terms to {self._get_initials(individual.full_name)}'))
            self.stdout.write(self.style.SUCCESS(f"Updated individual: {self._get_initials(individual.full_name)}"))
            sample_types = [s.strip() for s in (row_dict.get('Örnek Tipi') or '').split(',')]
            samples_touched = []
            for sample_type_name in sample_types:
                if not sample_type_name:
                    continue
                sample_type = self._get_or_create_sample_type(sample_type_name, admin_user)
                if not sample_type:
                    continue
                isolation_by = self._get_or_create_user(row_dict.get('İzolasyonu yapan'), admin_user)
                existing_sample = None
                if sample_type_name in ["Tam Kan", "Tam Kan/Serum"]:
                    existing_sample = Sample.objects.filter(
                        individual=individual,
                        sample_type__name__in=["Tam Kan", "Tam Kan/Serum"]
                    ).first()
                else:
                    existing_sample = Sample.objects.filter(
                        individual=individual,
                        sample_type=sample_type
                    ).first()
                if existing_sample:
                    if not existing_sample.receipt_date:
                        existing_sample.receipt_date = self._parse_date(row_dict.get('Geliş Tarihi/ay/gün/yıl'))
                        if existing_sample.receipt_date:
                            existing_sample.status = Status.objects.get(name='Available', content_type=ContentType.objects.get(app_label='lab', model='sample'))
                    if not existing_sample.isolation_by:
                        existing_sample.isolation_by = isolation_by
                    if not existing_sample.status:
                        existing_sample.status = Status.objects.get(name='Available', content_type=ContentType.objects.get(app_label='lab', model='sample'))
                    existing_sample.save()
                    self.stdout.write(self.style.SUCCESS(f'Updated existing sample: {existing_sample}'))
                    samples_touched.append(existing_sample)
                else:
                    initial_status = Status.objects.get(name='Available', content_type=ContentType.objects.get(app_label='lab', model='sample')) if self._parse_date(row_dict.get('Geliş Tarihi/ay/gün/yıl')) else Status.objects.get(name='Pending Blood Recovery', content_type=ContentType.objects.get(app_label='lab', model='sample'))
                    sample = Sample.objects.create(
                        individual=individual,
                        sample_type=sample_type,
                        status=initial_status,
                        receipt_date=self._parse_date(row_dict.get('Geliş Tarihi/ay/gün/yıl')),
                        isolation_by=isolation_by,
                        created_by=admin_user,
                    )
                    self.stdout.write(self.style.SUCCESS(f'Created new sample: {sample}'))
                    samples_touched.append(sample)
            # After touching samples for this row, add sample-level notes to all
            sample_notes = row_dict.get('Örnek Notları')
            if sample_notes and samples_touched:
                for s in samples_touched:
                    self._parse_and_add_notes(sample_notes, s, admin_user)
        self.stdout.write(self.style.WARNING(f'Leftover Rows: {len(leftover_rows)}'))
        if leftover_rows:
            with open(leftovers_path, 'w', encoding='utf-8', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=headers_lab, delimiter='\t')
                writer.writeheader()
                writer.writerows(leftover_rows)
            self.stdout.write(self.style.WARNING(f'Saved {len(leftover_rows)} rows with missing data to {leftovers_path}'))
        self.stdout.write(self.style.SUCCESS('Data import completed successfully'))
        # --- ANALIZ TAKIP SHEET ---
        ws_analiz = wb[analiz_takip_sheet]
        headers_analiz = [cell.value for cell in next(ws_analiz.iter_rows(min_row=1, max_row=1))]
        # Filter out None column names (empty columns at the end)
        headers_analiz = [h for h in headers_analiz if h is not None]
        self.stdout.write(f'Analiz Takip sheet headers: {headers_analiz}')
        
        gennext_type = self._get_or_create_analysis_type(
            'Gennext',
            '',
            admin_user
        )

        test_types = {}
        leftover_rows = []
        leftovers_path = os.path.join(os.path.dirname(file_path), 'import_analiz_takip_leftovers.tsv')

        # Track tests touched per individual (by lab_id) in this section to apply notes
        tests_touched_by_lab_id = {}

        for row in ws_analiz.iter_rows(min_row=2, values_only=True):
            # Skip rows where all values are None (empty rows at the end)
            if all(cell is None for cell in row):
                continue
            # Truncate row to match filtered headers (remove empty columns)
            row = row[:len(headers_analiz)]
            row_dict = dict(zip(headers_analiz, row))
            lab_id = row_dict.get('Özbek Lab. ID')
            if not lab_id:
                self.stdout.write(f'Skipping row with missing lab ID. Row data: {row_dict}')
                leftover_rows.append(row_dict)
                continue
            try:
                individual = Individual.objects.filter(cross_ids__id_value=lab_id).first()
                if not individual:
                    self.stdout.write(f'Individual not found with lab_id: {lab_id}. Available cross_ids: {list(Individual.objects.values_list("cross_ids__id_value", flat=True))[:10]}')
                    leftover_rows.append(row_dict)
                    continue
                # No default project assignment; projects handled via 'Projeler' in OZBEK LAB sheet
            except Exception as e:
                leftover_rows.append(row_dict)
                self.stdout.write(self.style.WARNING(f'Error finding individual for {lab_id}: {str(e)}'))
                continue
            veri_kaynagi = row_dict.get('VERİ KAYNAĞI')
            if not veri_kaynagi:
                self.stdout.write(f'Skipping row with missing VERİ KAYNAĞI for lab_id: {lab_id}. Available fields: {list(row_dict.keys())}')
                leftover_rows.append(row_dict)
                continue
            # Support multiple comma-separated test types in VERİ KAYNAĞI
            test_type_names = [p.strip() for p in str(veri_kaynagi).split(',') if p and str(p).strip()]
            if not test_type_names:
                leftover_rows.append(row_dict)
                self.stdout.write(self.style.WARNING(f'Skipping row with invalid/empty test type(s) for lab_id: {lab_id}'))
                continue
            sample = individual.samples.first()
            if not sample:
                self.stdout.write(self.style.WARNING(f'No sample found for individual: {self._get_initials(individual.full_name)}'))
                sample = self._get_or_create_placeholder_sample(individual, admin_user)
            created_tests_for_row = []
            for tt_name in test_type_names:
                test_type = self._get_or_create_test_type(tt_name, admin_user)
                if not test_type:
                    # Shouldn't happen due to guard, but keep parity with previous logic
                    self.stdout.write(self.style.WARNING(f'Skipping invalid test type "{tt_name}" for lab_id: {lab_id}'))
                    continue
                test_types[tt_name] = test_type
                test = Test.objects.create(
                    test_type=test_type,
                    status=Status.objects.get(name='Completed', content_type=ContentType.objects.get(app_label='lab', model='test')),
                    data_receipt_date=self._parse_date(row_dict.get('Data Geliş Tarihi')),
                    sample=sample,
                    created_by=admin_user
                )
                created_tests_for_row.append(test)
                # Track touched/created test for this individual's lab id
                tests_touched_by_lab_id.setdefault(lab_id, []).append(test)
            # Create Analysis records for each created test, if upload date provided
            data_upload_date = self._parse_date(row_dict.get('Data yüklenme tarihi/emre'))
            if data_upload_date and created_tests_for_row:
                for t in created_tests_for_row:
                    Analysis.objects.create(
                        type=gennext_type,
                        status=Status.objects.get(name='In Progress', content_type=ContentType.objects.get(app_label='lab', model='analysis')),
                        performed_date=data_upload_date,
                        performed_by=admin_user,
                        test=t,
                        created_by=admin_user
                    )
            # Add Test Notları for all tests touched for this individual
            test_notes = row_dict.get('Test Notları')
            if test_notes and tests_touched_by_lab_id.get(lab_id):
                for t in tests_touched_by_lab_id.get(lab_id, []):
                    self._parse_and_add_notes(test_notes, t, admin_user)
        if leftover_rows:
            with open(leftovers_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers_analiz, delimiter='\t')
                writer.writeheader()
                writer.writerows(leftover_rows)
            self.stdout.write(self.style.SUCCESS(f'Saved {len(leftover_rows)} rows to {leftovers_path}'))
        self.stdout.write(self.style.SUCCESS('Successfully imported analysis tracking data'))

        # --- SET PARENTS (mother/father) BASED ON RAREBOOST ID SUFFIXES ---
        # Convention: familyId.1 = proband, .2 = mother, .3 = father,
        #             .1.1, .1.2 = affected siblings. We set parents for 1 and 1.* members.
        try:
            updated_links = 0
            for fam_id, fam_obj in families.items():
                # Build code -> individual map using RareBoost ID
                code_to_individual = {}
                for ind in fam_obj.individuals.all():
                    lab_id_val = ind.lab_id
                    if not lab_id_val or lab_id_val == 'No Lab ID':
                        continue
                    parts = str(lab_id_val).split('.')
                    if len(parts) < 2:
                        continue
                    code_str = '.'.join(parts[1:])
                    # Prefer the first occurrence if duplicates exist
                    code_to_individual.setdefault(code_str, ind)

                mother_individual = code_to_individual.get('2')
                father_individual = code_to_individual.get('3')
                if not mother_individual and not father_individual:
                    continue

                # Assign parents for proband (1) and siblings (1.*)
                for code_str, child in code_to_individual.items():
                    if code_str == '1' or code_str.startswith('1.'):
                        changed = False
                        if mother_individual and child.mother_id != mother_individual.id and child.id != mother_individual.id:
                            child.mother = mother_individual
                            changed = True
                        if father_individual and child.father_id != father_individual.id and child.id != father_individual.id:
                            child.father = father_individual
                            changed = True
                        if changed:
                            child.save()
                            updated_links += 1
            if updated_links:
                self.stdout.write(self.style.SUCCESS(f"Parent links set/updated for {updated_links} individual(s) based on RareBoost IDs"))
            else:
                self.stdout.write("No parent links needed/updated based on RareBoost IDs")
        except Exception as e:
            self.stdout.write(self.style.WARNING(f"Error while setting parent links: {str(e)}"))

        # --- VARIANT IMPORT ---
        self.stdout.write("Starting Variant Import...")
        variant_sheet_name = None
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if headers and 'Chromosomal Position' in headers:
                variant_sheet_name = sheet_name
                break
        
        if not variant_sheet_name:
            self.stdout.write(self.style.WARNING("No sheet found with 'Chromosomal Position' header. Skipping variant import."))
        else:
            self.stdout.write(self.style.SUCCESS(f"Found variant sheet: {variant_sheet_name}"))
            ws_variant = wb[variant_sheet_name]
            headers_variant = [cell.value for cell in next(ws_variant.iter_rows(min_row=1, max_row=1))]
            # Do not filter None headers to preserve alignment
            
            for row in ws_variant.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue
                # row = row[:len(headers_variant)] # No need to truncate if we zip
                row_dict = dict(zip(headers_variant, row))
                
                lab_id = row_dict.get('Özbek Lab. ID')
                if not lab_id:
                    continue
                
                # Find Individual
                individual = Individual.objects.filter(cross_ids__id_value=lab_id).first()
                if not individual:
                    self.stdout.write(self.style.WARNING(f"Individual not found for variant import: {lab_id}"))
                    continue
                
                chrom_pos = row_dict.get('Chromosomal Position')
                parsed_variant = self._parse_variant_string(chrom_pos)
                if not parsed_variant:
                    self.stdout.write(self.style.WARNING(f"Could not parse chromosomal position: {chrom_pos}"))
                    continue
                
                chrom, start, ref, alt = parsed_variant
                
                # Create or Get SNV
                # We assume SNV for now based on the format
                try:
                    snv, created = SNV.objects.get_or_create(
                        individual=individual,
                        chromosome=chrom,
                        start=int(start),
                        end=int(start), # SNV start=end usually, or end=start+len(ref)-1? 
                                        # Standard VCF: POS is 1-based start. REF/ALT. 
                                        # Variant model has start/end. 
                                        # For SNV A>G, length is 1. start=end.
                                        # If deletion/insertion, length differs.
                                        # The regex expects A>G (single chars?) No, [ACGT]+
                        reference=ref,
                        alternate=alt,
                        defaults={
                            'created_by': admin_user,
                            'zygosity': self._map_zygosity(row_dict.get('Zygosity')),
                            'assembly_version': 'hg38' # Assumption
                        }
                    )
                    
                    # If not created, update zygosity if needed?
                    if not created:
                        new_zygosity = self._map_zygosity(row_dict.get('Zygosity'))
                        if snv.zygosity != new_zygosity:
                            snv.zygosity = new_zygosity
                            snv.save()
                            
                    # Add Clinical Association as Note
                    clin_assoc = row_dict.get('Clinical Association')
                    if clin_assoc:
                        self._parse_and_add_notes(clin_assoc, snv, admin_user)
                        
                    self.stdout.write(self.style.SUCCESS(f"Imported variant for {lab_id}: {chrom}:{start} {ref}>{alt}"))
                    
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"Error importing variant for {lab_id}: {e}"))

        self.stdout.write(self.style.SUCCESS("Variant import process completed."))
